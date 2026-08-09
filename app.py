import json
import sqlite3
import base64
import datetime
import hashlib
import html
import pathlib
import pandas as pd
import streamlit as st
import plotly.express as px
import plotly.graph_objects as go
from plotly.subplots import make_subplots
import requests
import numpy as np
import openpyxl
from scipy.spatial.distance import pdist, squareform
from sklearn.manifold import MDS
from zoneinfo import ZoneInfo
import pydeck as pdk

st.set_page_config(layout="wide", page_title="Garden Bird Dashboard", page_icon="🐦")

# ---- Styling ----
st.markdown("""
<style>
  @import url('https://fonts.googleapis.com/css2?family=Cabin:ital,wght@0,400;0,500;0,600;0,700;1,400&display=swap');

  :root {
    --bg:     #f5f3ee;
    --panel:  #ffffff;
    --text:   #1a2416;
    --muted:  #4a5c44;
    --border: rgba(26,36,22,0.11);
    --shadow: 0 4px 20px rgba(26,36,22,0.07);
    --radius: 14px;
    --accent: #3d6b44;
  }

  .stApp {
    background: var(--bg);
    font-family: 'Cabin', ui-sans-serif, system-ui, sans-serif !important;
    color: var(--text) !important;
    color-scheme: light !important;
  }
  /* Apply Cabin + force dark text on all text-bearing elements —
     NOT span (which breaks icon fonts like Material Symbols) */
  .stApp h1, .stApp h2, .stApp h3, .stApp h4, .stApp h5, .stApp h6,
  .stApp p, .stApp label, .stApp a, .stApp li,
  .stApp td, .stApp th,
  .stApp caption,
  .stApp input, .stApp textarea, .stApp select,
  .stApp .stRadio label {
    font-family: 'Cabin', ui-sans-serif, system-ui, sans-serif !important;
    color: var(--text) !important;
  }
  /* Force dark text on spans and divs but preserve their font-family
     (so icon fonts like Material Symbols keep rendering as icons) */
  .stApp span, .stApp div {
    color: var(--text) !important;
  }
  /* Restore white for sidebar toggle button icons */
  [data-testid="stSidebarCollapseButton"] *,
  [data-testid="stSidebarNavExpandButton"] * {
    color: #ffffff !important;
    fill: #ffffff !important;
  }
  /* Muted text for captions and small helper text */
  .stApp .stCaption, .stApp [data-testid="stCaptionContainer"] span,
  .stApp small {
    color: var(--muted) !important;
  }
  /* Ensure markdown text inherits correctly */
  .stApp [data-testid="stMarkdownContainer"] p,
  .stApp [data-testid="stMarkdownContainer"] li,
  .stApp [data-testid="stMarkdownContainer"] span {
    color: var(--text) !important;
  }

  .block-container {
    padding-top: 1.5rem;
    padding-bottom: 3rem;
    max-width: 1280px;
  }

  /* Sidebar */
  section[data-testid="stSidebar"] > div {
    background: #edeae0 !important;
    border-right: 1px solid var(--border) !important;
  }
  section[data-testid="stSidebar"] label,
  section[data-testid="stSidebar"] p,
  section[data-testid="stSidebar"] span {
    color: var(--muted) !important;
    font-size: 0.88rem;
    font-weight: 500;
  }

  /* Headings */
  h1 {
    color: var(--text) !important;
    letter-spacing: -0.03em;
    font-weight: 700;
  }
  h2, h3 {
    color: var(--text) !important;
    letter-spacing: -0.02em;
  }

  /* Sidebar navigation radio buttons */
  section[data-testid="stSidebar"] .stRadio > div {
    gap: 2px !important;
  }
  section[data-testid="stSidebar"] .stRadio label {
    padding: 8px 12px !important;
    border-radius: 10px !important;
    cursor: pointer !important;
    transition: background 0.15s ease !important;
  }
  section[data-testid="stSidebar"] .stRadio label:hover {
    background: rgba(61,107,68,0.08) !important;
  }
  section[data-testid="stSidebar"] .stRadio label[data-checked="true"],
  section[data-testid="stSidebar"] .stRadio div[role="radiogroup"] label:has(input:checked) {
    background: rgba(61,107,68,0.12) !important;
    color: var(--accent) !important;
    font-weight: 600 !important;
  }

  /* KPI metric cards */
  div[data-testid="stMetric"] {
    background: var(--panel) !important;
    border: 1px solid var(--border) !important;
    border-radius: var(--radius) !important;
    padding: 18px 22px !important;
    box-shadow: var(--shadow) !important;
  }
  div[data-testid="stMetric"] label {
    color: var(--muted) !important;
    font-weight: 600 !important;
    font-size: 0.78rem !important;
    text-transform: uppercase;
    letter-spacing: 0.06em;
  }
  div[data-testid="stMetric"] [data-testid="stMetricValue"] {
    color: var(--text) !important;
    font-weight: 700 !important;
  }

  /* Inputs */
  div[data-baseweb="select"] * ,
  div[data-baseweb="input"] * ,
  div[data-baseweb="textarea"] * {
    color: var(--text) !important;
  }
  div[data-baseweb="select"] > div,
  div[data-baseweb="input"] > div,
  div[data-baseweb="textarea"] > div {
    background: #ffffff !important;
    border-radius: 10px !important;
    border-color: var(--border) !important;
  }
  section[data-testid="stMain"] .stRadio [role="radiogroup"] {
    gap: 8px !important;
    flex-wrap: wrap !important;
  }
  section[data-testid="stMain"] .stRadio [role="radiogroup"] label {
    background: #ffffff !important;
    border: 1px solid rgba(26,36,22,0.14) !important;
    border-radius: 999px !important;
    box-shadow: 0 2px 8px rgba(26,36,22,0.04) !important;
    padding: 7px 13px !important;
    margin: 0 !important;
  }
  section[data-testid="stMain"] .stRadio [role="radiogroup"] label > div:first-child {
    display: none !important;
  }
  section[data-testid="stMain"] .stRadio [role="radiogroup"] label[data-checked="true"],
  section[data-testid="stMain"] .stRadio [role="radiogroup"] label:has(input:checked) {
    background: rgba(61,107,68,0.13) !important;
    border-color: rgba(61,107,68,0.55) !important;
  }
  section[data-testid="stMain"] .stRadio [role="radiogroup"] label p {
    font-size: 0.95rem !important;
    font-weight: 700 !important;
    line-height: 1.2 !important;
  }

  /* Plotly modebar */
  .js-plotly-plot .plotly .modebar { opacity: 0.2; }
  .js-plotly-plot:hover .plotly .modebar { opacity: 0.85; }

  hr { border-color: var(--border) !important; margin: 1rem 0 !important; }

  /* Header: transparent so it takes no visual space but keeps sidebar toggle in the DOM */
  header[data-testid="stHeader"] {
    background: var(--bg) !important;
    box-shadow: none !important;
  }
  /* Hide only the menu button inside the toolbar — NOT the toolbar container itself,
     which breaks the header flexbox layout and makes the expand button disappear */
  [data-testid="stMainMenu"] { display: none !important; }
  .stDecoration { display: none !important; }

  /* Birthday banner animation */
  @keyframes birthday-shimmer {
    0%   { background-position: 0% 50%; }
    50%  { background-position: 100% 50%; }
    100% { background-position: 0% 50%; }
  }
  .birthday-banner {
    font-size: 3rem;
    font-weight: 800;
    letter-spacing: 0.25em;
    text-align: center;
    background: linear-gradient(
      270deg, #3d6b44, #b89040, #8c5a70, #4a7090, #d4ac60, #5c8c5c, #c47a5a
    );
    background-size: 400% 400%;
    -webkit-background-clip: text;
    -webkit-text-fill-color: transparent;
    background-clip: text;
    animation: birthday-shimmer 4s ease infinite;
    padding: 1rem 0 0.5rem 0;
    margin-bottom: 0;
  }

  /* Garden news feed */
  .news-feed-title {
    color: var(--text) !important;
    font-size: 2.25rem !important;
    font-weight: 800 !important;
    letter-spacing: 0 !important;
    line-height: 1.1 !important;
    margin: 0 0 18px 0 !important;
  }
  .daily-date-panel {
    background: #ffffff !important;
    border: 1px solid rgba(26,36,22,0.12) !important;
    border-radius: 8px !important;
    box-shadow: 0 3px 14px rgba(26,36,22,0.05) !important;
    padding: 18px 20px !important;
    margin: 0 0 18px 0 !important;
  }
  .daily-date-title {
    color: var(--text) !important;
    font-size: 1.9rem !important;
    font-weight: 800 !important;
    letter-spacing: 0 !important;
    line-height: 1.12 !important;
  }
  .daily-controls-spacer {
    height: 4px !important;
  }
  .news-card {
    --news-accent: var(--accent);
    background: #ffffff !important;
    border: 1px solid rgba(26,36,22,0.12) !important;
    border-left: 5px solid var(--news-accent) !important;
    border-radius: 8px !important;
    box-shadow: 0 3px 14px rgba(26,36,22,0.06) !important;
    display: grid !important;
    grid-template-columns: minmax(240px, 0.42fr) minmax(320px, 1fr) !important;
    gap: 18px !important;
    align-items: baseline !important;
    padding: 12px 16px !important;
    margin: 0 0 8px 0 !important;
  }
  .news-headline {
    color: var(--text) !important;
    font-size: 1.06rem !important;
    font-weight: 800 !important;
    line-height: 1.3 !important;
  }
  .news-detail {
    color: var(--muted) !important;
    font-size: 0.98rem !important;
    font-weight: 500 !important;
    line-height: 1.35 !important;
  }
  @media (max-width: 700px) {
    .news-feed-title {
      font-size: 1.85rem !important;
      margin-bottom: 12px !important;
    }
    .daily-date-title {
      font-size: 1.45rem !important;
    }
    .news-card {
      grid-template-columns: 1fr !important;
      gap: 5px !important;
      padding: 15px 16px 14px 16px !important;
    }
    .news-headline {
      font-size: 1.08rem !important;
    }
  }

  /* ── Sidebar toggle buttons — visual only, don't touch font/icons ──── */
  [data-testid="stSidebarCollapseButton"],
  [data-testid="stSidebarNavExpandButton"] {
    visibility: visible !important;
    opacity: 1 !important;
    background: var(--accent) !important;
    color: #ffffff !important;
    border: none !important;
    border-radius: 8px !important;
    box-shadow: 0 2px 8px rgba(26,36,22,0.22) !important;
  }
  /* Force icon colour to white so it reads clearly against the green */
  [data-testid="stSidebarCollapseButton"] span,
  [data-testid="stSidebarCollapseButton"] svg {
    color: #ffffff !important;
    fill: #ffffff !important;
  }
  [data-testid="stSidebarCollapseButton"]:hover,
  [data-testid="stSidebarNavExpandButton"]:hover {
    background: #2d5233 !important;
  }
</style>
""", unsafe_allow_html=True)


# ---- Colour system ----
# A cohesive British-garden palette: foliage greens, earth tones, sky blues,
# harvest golds, mossy olives, hedgerow berries. Nothing electric or garish.
NATURE_PALETTE = [
    "#3d6b44",  # deep forest
    "#4a7090",  # lake blue
    "#b89040",  # harvest gold
    "#7a5c3d",  # dark bark
    "#6b7c4a",  # olive moss
    "#8c5a70",  # bramble berry
    "#5c8c5c",  # leaf green
    "#6a90b0",  # sky blue
    "#d4ac60",  # warm amber
    "#a07850",  # warm earth
    "#8c9c60",  # lichen
    "#4a5c70",  # dusk blue
    "#c47a5a",  # autumn terracotta
    "#7aaa6a",  # fresh growth
    "#8ab4c8",  # pale sky
    "#c4a07a",  # sandy loam
    "#a3c47a",  # spring sage
    "#607080",  # slate
    "#8c6b8c",  # heather
    "#90a890",  # soft sage
]

# Meaningful colours for UK conservation status
STATUS_COLORS = {
    "Green":            "#5c8c5c",
    "Amber":            "#d4ac60",
    "Red":              "#c47a5a",
    "Review Recording": "#8c9c8c",
    "Introduced":       "#4a7090",
    "Migrant":          "#6a90b0",
    "Scarce Migrant":   "#8ab4c8",
}

NEWS_CATEGORY_COLORS = {
    "Record": "#3d6b44",
    "Activity": "#b89040",
    "Arrival": "#6a90b0",
    "Seasonal timing": "#7aaa6a",
    "Species": "#6b7c4a",
    "Garden year": "#8c5a70",
    "Dawn chorus": "#c47a5a",
    "Comeback": "#4a7090",
    "Absence": "#607080",
    "Community mix": "#5c8c5c",
    "Time of day": "#d4ac60",
    "Weather": "#4a7090",
    "Data quality": "#8c6b8c",
    "Diversity": "#7aaa6a",
}

DIET_COLORS = {
    "Insectivore":  "#6a90b0",
    "Granivore":    "#d4ac60",
    "Omnivore":     "#5c8c5c",
    "Frugivore":    "#c47a5a",
    "Carnivore":    "#8b4c4c",
    "Piscivore":    "#4a7090",
    "Herbivore":    "#a3c47a",
    "Unclassified": "#8c9c8c",
}

PRIMARY   = "#3d6b44"  # deep forest — main single-series colour
SECONDARY = "#4a7090"  # lake blue
TERTIARY  = "#b89040"  # harvest gold

# Green-to-forest colorscale for heatmap
HEATMAP_SCALE = [
    [0.00, "#f5f3ee"],
    [0.20, "#c8dfa0"],
    [0.50, "#7aaa6a"],
    [0.75, "#5c8c5c"],
    [1.00, "#2d5233"],
]

MONTH_LABELS = {
    1: "Jan", 2: "Feb", 3: "Mar", 4: "Apr",
    5: "May", 6: "Jun", 7: "Jul", 8: "Aug",
    9: "Sep", 10: "Oct", 11: "Nov", 12: "Dec",
}

TIME_BUCKET_COLORS = {
    "Dawn (5–8)":       "#d4ac60",
    "Morning (8–12)":   "#5c8c5c",
    "Afternoon (12–17)":"#b89040",
    "Dusk (17–20)":     "#8c5a70",
    "Night (20–5)":     "#4a5c70",
}

SEASON_COLORS = {
    "Spring": "#7aaa6a",
    "Summer": "#d4ac60",
    "Autumn": "#c47a5a",
    "Winter": "#4a7090",
}


def status_color_map(statuses):
    """Build a color_discrete_map for a list of status strings."""
    cmap = {}
    fallback = [c for c in NATURE_PALETTE if c not in STATUS_COLORS.values()]
    fi = 0
    for s in statuses:
        if s in STATUS_COLORS:
            cmap[s] = STATUS_COLORS[s]
        else:
            cmap[s] = fallback[fi % len(fallback)]
            fi += 1
    return cmap


def _hex_to_rgb(hex_color: str) -> str:
    """Convert '#rrggbb' to 'r, g, b' string for use in rgba()."""
    h = hex_color.lstrip("#")
    return f"{int(h[0:2], 16)}, {int(h[2:4], 16)}, {int(h[4:6], 16)}"


def style_fig(fig):
    fig.update_layout(
        template="plotly_white",
        paper_bgcolor="#ffffff",
        plot_bgcolor="#fafaf8",
        font=dict(
            family="Cabin, ui-sans-serif, system-ui, sans-serif",
            color="#1a2416",
            size=13,
        ),
        title=dict(
            font=dict(size=19, color="#1a2416"),
            x=0.01,
            xanchor="left",
        ),
        legend=dict(
            font=dict(size=12, color="#1a2416"),
            title=dict(font=dict(size=12, color="#4a5c44")),
            bgcolor="rgba(255,255,255,0.92)",
            bordercolor="rgba(26,36,22,0.10)",
            borderwidth=1,
        ),
        margin=dict(l=10, r=10, t=58, b=10),
        hoverlabel=dict(
            bgcolor="#ffffff",
            bordercolor="rgba(26,36,22,0.15)",
            font=dict(size=13, color="#1a2416"),
        ),
    )
    fig.update_xaxes(
        showgrid=True,
        gridcolor="rgba(26,36,22,0.06)",
        zeroline=False,
        linecolor="rgba(26,36,22,0.10)",
        title_font=dict(size=13, color="#4a5c44"),
        tickfont=dict(size=12, color="#4a5c44"),
    )
    fig.update_yaxes(
        showgrid=True,
        gridcolor="rgba(26,36,22,0.06)",
        zeroline=False,
        linecolor="rgba(26,36,22,0.10)",
        title_font=dict(size=13, color="#4a5c44"),
        tickfont=dict(size=12, color="#4a5c44"),
    )
    return fig


def assign_time_bucket(hour):
    if 5 <= hour < 8:
        return "Dawn (5–8)"
    elif 8 <= hour < 12:
        return "Morning (8–12)"
    elif 12 <= hour < 17:
        return "Afternoon (12–17)"
    elif 17 <= hour < 20:
        return "Dusk (17–20)"
    else:
        return "Night (20–5)"


@st.cache_data
def compute_nmds(feature_matrix, species_list):
    dist = squareform(pdist(feature_matrix, metric="braycurtis"))
    mds = MDS(
        n_components=2,
        metric=False,
        dissimilarity="precomputed",
        n_init=10,
        max_iter=500,
        random_state=42,
    )
    coords = mds.fit_transform(dist)
    stress = mds.stress_
    return coords, stress


@st.cache_data(ttl=86400)
def fetch_wiki_summary(title: str):
    """Fetch a Wikipedia summary for the given page title."""
    url = f"https://en.wikipedia.org/api/rest_v1/page/summary/{title}"
    headers = {"User-Agent": "GardenBirdDashboard/1.0 (https://github.com/emjgood1995/bird-dashboard)"}
    try:
        resp = requests.get(url, headers=headers, timeout=10)
        if resp.status_code != 200:
            return None
        data = resp.json()
        return {
            "extract": data.get("extract", ""),
            "thumbnail_url": data.get("thumbnail", {}).get("source"),
            "page_url": data.get("content_urls", {}).get("desktop", {}).get("page", ""),
            "title": data.get("title", title),
        }
    except Exception:
        return None


@st.cache_data(ttl=86400)
def fetch_bird_audio(sci_name: str):
    """Fetch a bird song/call recording from Wikimedia Commons."""
    api = "https://commons.wikimedia.org/w/api.php"
    headers = {"User-Agent": "GardenBirdDashboard/1.0 (https://github.com/emjgood1995/bird-dashboard)"}
    try:
        # Search for audio files matching the scientific name
        resp = requests.get(api, params={
            "action": "query", "list": "search", "format": "json",
            "srsearch": f"{sci_name} song filetype:audio",
            "srnamespace": "6", "srlimit": "1",
        }, headers=headers, timeout=10)
        if resp.status_code != 200:
            return None
        results = resp.json().get("query", {}).get("search", [])
        if not results:
            # Fallback: search without "song"
            resp = requests.get(api, params={
                "action": "query", "list": "search", "format": "json",
                "srsearch": f"{sci_name} filetype:audio",
                "srnamespace": "6", "srlimit": "1",
            }, headers=headers, timeout=10)
            if resp.status_code != 200:
                return None
            results = resp.json().get("query", {}).get("search", [])
        if not results:
            return None
        title = results[0]["title"]
        # Get the direct file URL
        resp = requests.get(api, params={
            "action": "query", "titles": title, "format": "json",
            "prop": "imageinfo", "iiprop": "url|mime",
        }, headers=headers, timeout=10)
        if resp.status_code != 200:
            return None
        pages = resp.json().get("query", {}).get("pages", {})
        for page in pages.values():
            info = page.get("imageinfo", [{}])[0]
            file_url = info.get("url", "")
            desc_url = info.get("descriptionurl", "")
            mime = info.get("mime", "")
            if file_url:
                fmt = "audio/ogg" if "ogg" in mime else "audio/mpeg"
                return {"file": file_url, "format": fmt, "page": desc_url, "title": title}
        return None
    except Exception:
        return None


# ---- Load data ----
DB_PATH = pathlib.Path("birds_lfs.db")


def database_cache_signature(db_path):
    stat = db_path.stat()
    return str(db_path), stat.st_mtime_ns, stat.st_size


@st.cache_data(max_entries=2)
def load_data(db_path, db_mtime_ns, db_size):
    # The mtime and size arguments make Streamlit reload when cron updates the DB.
    conn = sqlite3.connect(db_path)
    df = pd.read_sql_query("SELECT * FROM detections", conn)
    conn.close()

    df["timestamp"] = pd.to_datetime(df["Date"] + " " + df["Time"], errors="coerce")
    df["hour"]  = df["timestamp"].dt.hour
    df["week"]  = df["timestamp"].dt.isocalendar().week.astype(int)
    df["month"] = df["timestamp"].dt.month.astype(int)

    meta = pd.read_excel("UK_Birds_Generalized_Status.xlsx")
    meta = meta.rename(columns={
        "Latin Name":  "Sci_Name",
        "Common Name": "UK_Common_Name",
        "Status":      "UK_Status",
    })
    meta = meta[["Sci_Name", "UK_Common_Name", "UK_Status"]].drop_duplicates()

    df = df.merge(meta, on="Sci_Name", how="left")
    df["UK_Status"] = df["UK_Status"].fillna("Review Recording")
    return df

df = load_data(*database_cache_signature(DB_PATH))


def load_diet_map():
    try:
        with open("species_diet.json") as f:
            return json.load(f)
    except FileNotFoundError:
        return {}

_diet_map = load_diet_map()
df["Diet"] = df["Sci_Name"].map(_diet_map).fillna("Unclassified")


@st.cache_data(ttl=86400)
def fetch_weather(lat: float, lon: float, start_date: str, end_date: str):
    """Fetch historical hourly weather from Open-Meteo and return a DataFrame."""
    url = (
        f"https://archive-api.open-meteo.com/v1/archive"
        f"?latitude={lat}&longitude={lon}"
        f"&start_date={start_date}&end_date={end_date}"
        f"&hourly=temperature_2m,precipitation,wind_speed_10m,cloud_cover,pressure_msl"
        f"&daily=temperature_2m_max,temperature_2m_min,precipitation_sum,wind_speed_10m_max,sunrise,sunset"
        f"&timezone=Europe%2FLondon"
    )
    try:
        resp = requests.get(url, timeout=30)
        if resp.status_code != 200:
            return None, None
        data = resp.json()

        hourly = pd.DataFrame({
            "datetime": pd.to_datetime(data["hourly"]["time"]),
            "temperature": data["hourly"]["temperature_2m"],
            "precipitation": data["hourly"]["precipitation"],
            "wind_speed": data["hourly"]["wind_speed_10m"],
            "cloud_cover": data["hourly"]["cloud_cover"],
            "pressure": data["hourly"]["pressure_msl"],
        })
        hourly["date"] = hourly["datetime"].dt.date
        hourly["hour"] = hourly["datetime"].dt.hour

        # Parse sunrise/sunset as Europe/London aware, then convert to UTC
        _tz_london = ZoneInfo("Europe/London")
        _tz_utc = ZoneInfo("UTC")
        sunrise_local = pd.to_datetime(data["daily"]["sunrise"])
        sunset_local = pd.to_datetime(data["daily"]["sunset"])
        sunrise_utc = sunrise_local.map(lambda t: t.replace(tzinfo=_tz_london).astimezone(_tz_utc).replace(tzinfo=None))
        sunset_utc = sunset_local.map(lambda t: t.replace(tzinfo=_tz_london).astimezone(_tz_utc).replace(tzinfo=None))

        daily = pd.DataFrame({
            "date": pd.to_datetime(data["daily"]["time"]).date,
            "temp_max": data["daily"]["temperature_2m_max"],
            "temp_min": data["daily"]["temperature_2m_min"],
            "precip_sum": data["daily"]["precipitation_sum"],
            "wind_max": data["daily"]["wind_speed_10m_max"],
            "sunrise": sunrise_local,
            "sunset": sunset_local,
            "sunrise_utc": sunrise_utc,
            "sunset_utc": sunset_utc,
        })
        return hourly, daily
    except Exception:
        return None, None


@st.cache_data(ttl=3600)
def fetch_inat_nearby(lat, lon, radius_km=25, days_back=30, per_page=200):
    """Fetch recent bird observations from iNaturalist near a location."""
    d2 = datetime.date.today()
    d1 = d2 - datetime.timedelta(days=days_back)
    url = "https://api.inaturalist.org/v1/observations"
    params = {
        "lat": lat, "lng": lon, "radius": radius_km,
        "iconic_taxa": "Aves",
        "quality_grade": "research",
        "d1": d1.isoformat(), "d2": d2.isoformat(),
        "per_page": per_page,
        "order_by": "observed_on",
        "fields": "taxon,location,observed_on,photos,place_guess,uri",
    }
    try:
        resp = requests.get(url, params=params, timeout=30,
                            headers={"User-Agent": "GardenBirdDashboard/1.0"})
        if resp.status_code != 200:
            return None
        return resp.json()
    except Exception:
        return None


_TZ_LONDON = ZoneInfo("Europe/London")
_TZ_UTC = ZoneInfo("UTC")


def to_utc_hour(ts: pd.Series) -> pd.Series:
    """Convert naive local (Europe/London) detection timestamps to UTC decimal hours."""
    utc_ts = ts.apply(lambda t: t.replace(tzinfo=_TZ_LONDON).astimezone(_TZ_UTC) if pd.notna(t) else t)
    return utc_ts.dt.hour + utc_ts.dt.minute / 60.0


DAILY_PERIOD_OPTIONS = ["Day", "Last 7 days", "Last 30 days"]
VISIBLE_HEADLINE_LIMIT = 10
NEWS_CHART_HEIGHT = 200
GARDEN_EVENTS_PATH = pathlib.Path("garden_events.json")


def daily_period_bounds(end_date, period_mode):
    if end_date is None:
        return None, None
    if period_mode == "Last 7 days":
        return end_date - datetime.timedelta(days=6), end_date
    if period_mode == "Last 30 days":
        return end_date - datetime.timedelta(days=29), end_date
    return end_date, end_date


def format_period_label(start_date, end_date):
    if start_date is None or end_date is None:
        return "No dates available"
    if start_date == end_date:
        return end_date.strftime("%A %d %B %Y")
    if start_date.year == end_date.year:
        return f"{start_date.strftime('%d %b')} to {end_date.strftime('%d %b %Y')}"
    return f"{start_date.strftime('%d %b %Y')} to {end_date.strftime('%d %b %Y')}"


def filter_date_window(data, start_date, end_date):
    if start_date is None or end_date is None or len(data) == 0:
        return data.iloc[0:0].copy()
    return data[
        (data["timestamp"].dt.date >= start_date) &
        (data["timestamp"].dt.date <= end_date)
    ].copy()


@st.cache_data
def load_garden_events():
    try:
        with open(GARDEN_EVENTS_PATH) as f:
            events = json.load(f)
        return events if isinstance(events, list) else []
    except (FileNotFoundError, json.JSONDecodeError):
        return []


def prepare_news_df(data):
    news_df = data.dropna(subset=["timestamp"]).copy()
    if len(news_df) == 0:
        return news_df
    news_df["date"] = news_df["timestamp"].dt.date
    news_df["hour"] = news_df["timestamp"].dt.hour.astype(int)
    news_df["year"] = news_df["timestamp"].dt.year.astype(int)
    news_df["month_num"] = news_df["timestamp"].dt.month.astype(int)
    news_df["doy"] = news_df["timestamp"].dt.dayofyear.astype(int)
    return news_df


def add_insight(insights, priority, category, headline, detail, species=None, chart=None):
    insight = {
        "priority": priority,
        "category": category,
        "headline": headline,
        "detail": detail,
        "species": species,
    }
    if chart is not None:
        insight["chart"] = chart
    insights.append(insight)


def insight_key(insight):
    return (insight["category"], insight["headline"], insight.get("species"))


def pct_change(current, baseline):
    if baseline is None or baseline == 0 or pd.isna(baseline):
        return None
    return ((current - baseline) / baseline) * 100


def signed_pct_label(value):
    if value is None or pd.isna(value):
        return ""
    sign = "+" if value > 0 else ""
    return f"{sign}{value:.0f}%"


def period_noun(period_days):
    if period_days == 1:
        return "day"
    return f"{period_days}-day period"


def news_season_from_month(month):
    if month in (3, 4, 5):
        return "Spring"
    if month in (6, 7, 8):
        return "Summer"
    if month in (9, 10, 11):
        return "Autumn"
    return "Winter"


def date_label(date_value):
    if pd.isna(date_value):
        return "unknown date"
    if hasattr(date_value, "date"):
        date_value = date_value.date()
    return date_value.strftime("%d %b %Y")


def doy_label(doy):
    if pd.isna(doy):
        return None
    ref = datetime.date(2000, 1, 1) + datetime.timedelta(days=int(round(doy)) - 1)
    return f"{ref.day} {ref.strftime('%b')}"


def event_species_mask(data, event):
    names = data["Com_Name"].fillna("").astype(str)
    names_lower = names.str.lower()
    mask = pd.Series(False, index=data.index)

    exact_names = {str(s).lower() for s in event.get("species", [])}
    if exact_names:
        mask = mask | names_lower.isin(exact_names)

    for term in event.get("match_terms", []):
        term = str(term).strip().lower()
        if term:
            mask = mask | names_lower.str.contains(term, regex=False)

    return mask


def matching_events(species, events, trigger=None, months=None):
    species_lower = str(species).lower()
    matched = []
    for event in events:
        if trigger is not None and event.get("trigger") != trigger:
            continue
        event_months = set(event.get("months", []))
        if months is not None and event_months and not event_months.intersection(months):
            continue

        exact_names = {str(s).lower() for s in event.get("species", [])}
        term_match = any(
            str(term).strip().lower() in species_lower
            for term in event.get("match_terms", [])
            if str(term).strip()
        )
        if species_lower in exact_names or term_match:
            matched.append(event)
    return matched


def comparison_window(all_news, current_news, start_date, end_date, prefer_same_month=True):
    if len(all_news) == 0:
        return all_news.iloc[0:0].copy()

    period_days = (end_date - start_date).days + 1
    not_current = ~((all_news["date"] >= start_date) & (all_news["date"] <= end_date))
    comparison = all_news[not_current].copy()

    if prefer_same_month and len(current_news):
        months = set(current_news["month_num"].dropna().astype(int).unique().tolist())
        same_month = comparison[comparison["month_num"].isin(months)].copy()
        if same_month["date"].nunique() >= max(7, period_days):
            return same_month

    return comparison


def expected_count_for_period(all_news, current_news, start_date, end_date, filter_mask=None):
    period_days = (end_date - start_date).days + 1
    comp = comparison_window(all_news, current_news, start_date, end_date)
    if filter_mask is not None and len(comp):
        comp = comp[filter_mask(comp)].copy()
    comp_days = comp["date"].nunique()
    if comp_days == 0:
        return None
    return (len(comp) / comp_days) * period_days


def decimal_hour(ts):
    return ts.dt.hour + ts.dt.minute / 60.0


def hour_text(hour_value):
    hour_int = int(hour_value)
    minute_int = int(round((hour_value - hour_int) * 60))
    if minute_int == 60:
        hour_int += 1
        minute_int = 0
    return f"{hour_int:02d}:{minute_int:02d}"


def minutes_text(minutes):
    minutes = int(round(abs(minutes)))
    if minutes < 60:
        return f"{minutes} minutes"
    hours = minutes // 60
    remainder = minutes % 60
    if remainder == 0:
        return f"{hours} hour{'s' if hours != 1 else ''}"
    return f"{hours}h {remainder}m"


def longest_date_streak(dates):
    unique_days = sorted(set(dates))
    if not unique_days:
        return 0
    best = 1
    current = 1
    for i in range(1, len(unique_days)):
        if (unique_days[i] - unique_days[i - 1]).days == 1:
            current += 1
            best = max(best, current)
        else:
            current = 1
    return best


def streak_ending_at(dates, end_date):
    unique_days = sorted(set(d for d in dates if d <= end_date))
    if not unique_days or unique_days[-1] != end_date:
        return 0
    streak = 1
    for i in range(len(unique_days) - 1, 0, -1):
        if (unique_days[i] - unique_days[i - 1]).days == 1:
            streak += 1
        else:
            break
    return streak


def daily_species_sets(data):
    if len(data) == 0:
        return {}
    return data.groupby("date")["Com_Name"].apply(lambda x: frozenset(x.dropna().astype(str))).to_dict()


def species_mix_similarity(current_counts, comparison_counts):
    species = sorted(set(current_counts.index).union(set(comparison_counts.index)))
    if not species:
        return 0
    current_vec = np.array([current_counts.get(sp, 0) for sp in species], dtype=float)
    comparison_vec = np.array([comparison_counts.get(sp, 0) for sp in species], dtype=float)
    denom = np.linalg.norm(current_vec) * np.linalg.norm(comparison_vec)
    if denom == 0:
        return 0
    return float(np.dot(current_vec, comparison_vec) / denom)


def add_period_record_insights(all_news, current_news, start_date, end_date, insights):
    if len(all_news) == 0 or len(current_news) == 0:
        return

    period_days = (end_date - start_date).days + 1
    noun = period_noun(period_days)
    current_count = len(current_news)
    current_species = current_news["Com_Name"].nunique()

    all_dates = pd.date_range(all_news["date"].min(), all_news["date"].max(), freq="D").date
    daily_counts = all_news.groupby("date").size().reindex(all_dates, fill_value=0)
    rolling_counts = daily_counts.rolling(period_days, min_periods=period_days).sum().dropna()

    if len(rolling_counts) >= 5:
        max_count = int(rolling_counts.max())
        median_count = float(rolling_counts.median())
        before_current = rolling_counts[rolling_counts.index < end_date]
        if len(before_current) and current_count >= max_count and current_count > 0:
            add_insight(
                insights,
                95,
                "Record",
                f"Busiest {noun} on record",
                f"{current_count:,} detections, ahead of the previous high of {int(before_current.max()):,}.",
                chart={"type": "activity_period", "metric": "detections"},
            )
        elif median_count > 0:
            change = pct_change(current_count, median_count)
            if change is not None and change >= 45 and current_count - median_count >= max(25, median_count * 0.35):
                add_insight(
                    insights,
                    78,
                    "Activity",
                    f"Unusually busy {noun}",
                    f"{current_count:,} detections, {signed_pct_label(change)} vs the typical {noun}.",
                    chart={"type": "activity_period", "metric": "detections"},
                )
            elif change is not None and change <= -45 and median_count - current_count >= max(20, median_count * 0.35):
                add_insight(
                    insights,
                    74,
                    "Activity",
                    f"Quiet {noun}",
                    f"{current_count:,} detections, {signed_pct_label(change)} vs the typical {noun}.",
                    chart={"type": "activity_period", "metric": "detections"},
                )

    if period_days == 1:
        daily_species = all_news.groupby("date")["Com_Name"].nunique().reindex(all_dates, fill_value=0)
        if len(daily_species) >= 5 and current_species >= int(daily_species.max()) and current_species > 0:
            add_insight(
                insights,
                86,
                "Diversity",
                "Highest species count for a day",
                f"{current_species:,} species recorded on {date_label(end_date)}.",
                chart={"type": "species_mix_period"},
            )
        elif len(daily_species) >= 5 and daily_species.median() > 0:
            species_change = pct_change(current_species, daily_species.median())
            if species_change is not None and species_change >= 45:
                add_insight(
                    insights,
                    68,
                    "Diversity",
                    "Species mix was richer than usual",
                    f"{current_species:,} species, {signed_pct_label(species_change)} vs a typical day.",
                    chart={"type": "species_mix_period"},
                )


def add_arrival_insights(all_news, current_news, start_date, end_date, events, insights):
    if len(all_news) == 0 or len(current_news) == 0:
        return

    year = end_date.year
    year_df = all_news[all_news["year"] == year].copy()
    if len(year_df) == 0:
        return

    first_seen = (
        year_df.groupby("Com_Name")["timestamp"]
        .min()
        .reset_index(name="First_Seen")
    )
    first_seen["First_Date"] = first_seen["First_Seen"].dt.date
    arrivals = first_seen[
        (first_seen["First_Date"] >= start_date) &
        (first_seen["First_Date"] <= end_date)
    ].copy()

    if len(arrivals) == 0:
        return

    prev_firsts = (
        all_news[all_news["year"] < year]
        .groupby(["year", "Com_Name"])["timestamp"]
        .min()
        .reset_index(name="First_Seen")
    )
    if len(prev_firsts):
        prev_firsts["doy"] = prev_firsts["First_Seen"].dt.dayofyear

    candidate_rows = []
    months = set(current_news["month_num"].dropna().astype(int).unique().tolist())
    for _, row in arrivals.iterrows():
        species = row["Com_Name"]
        event_matches = matching_events(species, events, trigger="first_seen_year", months=months)
        if start_date.month == 1 and not event_matches:
            continue

        detail = f"First detected this year on {date_label(row['First_Date'])}."
        timing_shift = None
        if len(prev_firsts):
            species_prev = prev_firsts[prev_firsts["Com_Name"] == species]
            if len(species_prev):
                typical_doy = species_prev["doy"].median()
                typical = doy_label(typical_doy)
                if typical:
                    detail = f"First detected this year on {date_label(row['First_Date'])}; typical first date is around {typical}."
                current_doy = int(row["First_Seen"].dayofyear)
                diff_days = current_doy - typical_doy
                if abs(diff_days) >= 10:
                    timing_shift = int(round(diff_days))

        if timing_shift is not None:
            direction = "later" if timing_shift > 0 else "earlier"
            days = abs(timing_shift)
            if event_matches:
                headline = f"{event_matches[0].get('label', species)} is {days} days {direction} than usual"
            else:
                headline = f"{species} arrived {days} days {direction} than usual"
            priority = 91 if event_matches else 78
        elif event_matches:
            headline = event_matches[0].get("label", f"{species} arrived")
            priority = 88
        else:
            headline = f"{species} first seen this year"
            priority = 62
        candidate_rows.append((priority, headline, detail, species))

    for priority, headline, detail, species in sorted(candidate_rows, reverse=True)[:4]:
        add_insight(
            insights,
            priority,
            "Arrival",
            headline,
            detail,
            species,
            chart={"type": "species_recent", "species": species},
        )


def add_species_change_insights(all_news, current_news, start_date, end_date, events, insights):
    if len(all_news) == 0 or len(current_news) == 0:
        return

    period_days = (end_date - start_date).days + 1
    months = set(current_news["month_num"].dropna().astype(int).unique().tolist())
    current_counts = current_news["Com_Name"].value_counts()

    comparison = all_news[
        ~((all_news["date"] >= start_date) & (all_news["date"] <= end_date)) &
        (all_news["month_num"].isin(months))
    ].copy()
    comparison_days = comparison["date"].nunique()
    if comparison_days < max(7, period_days):
        comparison = all_news[
            ~((all_news["date"] >= start_date) & (all_news["date"] <= end_date))
        ].copy()
        comparison_days = comparison["date"].nunique()
    if comparison_days == 0:
        return

    baseline_counts = comparison["Com_Name"].value_counts()
    species = sorted(set(current_counts.index).union(set(baseline_counts.index)))

    spikes = []
    drops = []
    for sp in species:
        expected = (baseline_counts.get(sp, 0) / comparison_days) * period_days
        current = int(current_counts.get(sp, 0))
        if expected >= 5 and current >= max(expected * 2.5, expected + 10):
            event_matches = matching_events(sp, events, trigger="spike", months=months)
            change = pct_change(current, expected)
            headline = event_matches[0].get("label", f"{sp} activity spiked") if event_matches else f"{sp} activity spiked"
            spikes.append((
                84 if event_matches else 70,
                headline,
                f"{current:,} detections vs about {expected:.0f} expected for this period ({signed_pct_label(change)}).",
                sp,
                {"type": "species_recent", "species": sp},
            ))
        elif expected >= 10 and current <= expected * 0.35:
            event_matches = matching_events(sp, events, trigger="drop_vs_last_year", months=months)
            change = pct_change(current, expected)
            headline = event_matches[0].get("label", f"{sp} unusually quiet") if event_matches else f"{sp} unusually quiet"
            drops.append((
                82 if event_matches else 68,
                headline,
                f"{current:,} detections vs about {expected:.0f} expected for this period ({signed_pct_label(change)}).",
                sp,
                {"type": "species_recent", "species": sp},
            ))

    for priority, headline, detail, species_name, chart in sorted(spikes, key=lambda row: row[0], reverse=True)[:3]:
        add_insight(insights, priority, "Species", headline, detail, species_name, chart=chart)
    for priority, headline, detail, species_name, chart in sorted(drops, key=lambda row: row[0], reverse=True)[:3]:
        add_insight(insights, priority, "Species", headline, detail, species_name, chart=chart)


def add_event_ytd_insights(all_news, end_date, events, insights):
    if len(all_news) == 0:
        return

    year = end_date.year
    prev_year = year - 1
    year_start = datetime.date(year, 1, 1)
    prev_start = datetime.date(prev_year, 1, 1)

    try:
        prev_end = end_date.replace(year=prev_year)
    except ValueError:
        prev_end = datetime.date(prev_year, 2, 28)

    for event in events:
        if event.get("trigger") != "drop_vs_last_year":
            continue

        current_year = all_news[
            (all_news["date"] >= year_start) &
            (all_news["date"] <= end_date)
        ].copy()
        prev_year_df = all_news[
            (all_news["date"] >= prev_start) &
            (all_news["date"] <= prev_end)
        ].copy()

        current_count = int(event_species_mask(current_year, event).sum())
        previous_count = int(event_species_mask(prev_year_df, event).sum())
        if previous_count < 20:
            continue

        change = pct_change(current_count, previous_count)
        if change is not None and change <= -55:
            add_insight(
                insights,
                90,
                "Garden year",
                event.get("label", "A regular garden visitor is down"),
                (
                    f"{current_count:,} detections so far in {year}, compared with "
                    f"{previous_count:,} by {date_label(prev_end)} last year ({signed_pct_label(change)})."
                ),
            )


def add_dawn_chorus_insights(all_news, current_news, start_date, end_date, weather_daily, insights):
    if len(all_news) == 0 or len(current_news) == 0:
        return

    period_days = (end_date - start_date).days + 1
    dawn_chart = {"type": "hourly_activity", "highlight_hours": list(range(3, 11))}
    dawn = current_news[(current_news["hour"] >= 3) & (current_news["hour"] <= 10)].copy()
    if len(dawn) == 0:
        return

    comp = comparison_window(all_news, current_news, start_date, end_date)
    comp_dawn = comp[(comp["hour"] >= 3) & (comp["hour"] <= 10)].copy()
    comp_days = comp["date"].nunique()
    expected = (len(comp_dawn) / comp_days) * period_days if comp_days else None

    if expected is not None and expected >= 5:
        change = pct_change(len(dawn), expected)
        if len(dawn) >= max(expected * 1.8, expected + 15):
            add_insight(
                insights,
                76,
                "Dawn chorus",
                "Dawn chorus was busier than usual",
                f"{len(dawn):,} detections between 03:00 and 10:00, {signed_pct_label(change)} vs expected.",
                chart=dawn_chart,
            )
        elif len(dawn) <= expected * 0.45:
            add_insight(
                insights,
                70,
                "Dawn chorus",
                "Dawn chorus was unusually quiet",
                f"{len(dawn):,} detections between 03:00 and 10:00, {signed_pct_label(change)} vs expected.",
                chart=dawn_chart,
            )

    top_dawn = dawn["Com_Name"].value_counts()
    if len(top_dawn):
        top_species = top_dawn.index[0]
        top_count = int(top_dawn.iloc[0])
        top_share = top_count / len(dawn)
        if top_count >= 10 and top_share >= 0.45:
            add_insight(
                insights,
                73,
                "Dawn chorus",
                f"{top_species} dominated the dawn window",
                f"{top_count:,} of {len(dawn):,} dawn detections ({top_share:.0%}) were {top_species}.",
                top_species,
                chart=dawn_chart,
            )

    if period_days == 1 and len(comp_dawn):
        current_first = decimal_hour(dawn["timestamp"]).min()
        comp_first = comp_dawn.groupby("date")["timestamp"].min()
        if len(comp_first) >= 7:
            typical_first = decimal_hour(comp_first).median()
            diff_minutes = (current_first - typical_first) * 60
            if diff_minutes <= -30:
                add_insight(
                    insights,
                    80,
                    "Dawn chorus",
                    "Dawn chorus started earlier than usual",
                    f"First dawn detection was at {hour_text(current_first)}, {minutes_text(diff_minutes)} earlier than typical.",
                    chart=dawn_chart,
                )
            elif diff_minutes >= 45:
                add_insight(
                    insights,
                    72,
                    "Dawn chorus",
                    "Dawn chorus started later than usual",
                    f"First dawn detection was at {hour_text(current_first)}, {minutes_text(diff_minutes)} later than typical.",
                    chart=dawn_chart,
                )

        if weather_daily is not None and len(weather_daily):
            weather_match = weather_daily[weather_daily["date"] == end_date]
            if len(weather_match) and pd.notna(weather_match.iloc[0].get("sunrise")):
                sunrise = weather_match.iloc[0]["sunrise"]
                sunrise_hour = sunrise.hour + sunrise.minute / 60.0
                if current_first <= sunrise_hour - 0.5:
                    add_insight(
                        insights,
                        58,
                        "Dawn chorus",
                        "First detection came well before sunrise",
                        f"First dawn detection was at {hour_text(current_first)}; sunrise was around {hour_text(sunrise_hour)}.",
                        chart=dawn_chart,
                    )


def add_expected_arrival_insights(all_news, current_news, start_date, end_date, insights):
    if len(all_news) == 0:
        return

    year = end_date.year
    current_year = all_news[
        (all_news["year"] == year) &
        (all_news["date"] <= end_date)
    ].copy()
    previous_years = all_news[all_news["year"] < year].copy()
    if len(previous_years) == 0:
        return

    prev_firsts = (
        previous_years.groupby(["year", "Com_Name"])["timestamp"]
        .min()
        .reset_index(name="First_Seen")
    )
    if len(prev_firsts) == 0:
        return

    prev_firsts["doy"] = prev_firsts["First_Seen"].dt.dayofyear
    prev_counts = previous_years["Com_Name"].value_counts()
    current_seen = set(current_year["Com_Name"].dropna().unique())
    current_doy = end_date.timetuple().tm_yday

    expected = (
        prev_firsts.groupby("Com_Name")
        .agg(years_seen=("year", "nunique"), median_doy=("doy", "median"))
        .reset_index()
    )
    expected["prev_count"] = expected["Com_Name"].map(prev_counts).fillna(0)
    expected = expected[
        (expected["years_seen"] >= 2) &
        (expected["prev_count"] >= 10) &
        (expected["median_doy"] <= current_doy - 14) &
        (~expected["Com_Name"].isin(current_seen))
    ].copy()

    if len(expected) == 0:
        return

    expected["days_late"] = current_doy - expected["median_doy"]
    expected = expected.sort_values(["years_seen", "days_late", "prev_count"], ascending=False).head(3)
    for _, row in expected.iterrows():
        species = row["Com_Name"]
        add_insight(
            insights,
            67,
            "Seasonal timing",
            f"{species} has not appeared yet this year",
            f"Usually first detected around {doy_label(row['median_doy'])}; currently about {int(row['days_late'])} days later than that.",
            species,
            chart={"type": "species_recent", "species": species},
        )


def add_absence_comeback_insights(all_news, current_news, start_date, end_date, insights):
    if len(all_news) == 0:
        return

    period_days = (end_date - start_date).days + 1
    current_counts = current_news["Com_Name"].value_counts()
    before = all_news[all_news["date"] < start_date].copy()

    comebacks = []
    for species, count in current_counts.items():
        species_before = before[before["Com_Name"] == species]
        if len(species_before) == 0:
            continue
        last_before = species_before["date"].max()
        gap_days = (start_date - last_before).days - 1
        if gap_days >= max(7, period_days * 2) and count >= 2:
            comebacks.append((gap_days, int(count), species))

    for gap_days, count, species in sorted(comebacks, reverse=True)[:3]:
        add_insight(
            insights,
            75,
            "Comeback",
            f"{species} returned after a quiet spell",
            f"{count:,} detections after {gap_days} days without a detection.",
            species,
            chart={"type": "species_recent", "species": species},
        )

    lookback_days = max(30, period_days * 3)
    recent_start = start_date - datetime.timedelta(days=lookback_days)
    recent = all_news[(all_news["date"] >= recent_start) & (all_news["date"] < start_date)].copy()
    if len(recent) == 0:
        return

    recent_counts = recent["Com_Name"].value_counts()
    absences = []
    for species, recent_count in recent_counts.items():
        if species in current_counts:
            continue
        expected = (recent_count / max(recent["date"].nunique(), 1)) * period_days
        if recent_count >= 15 and expected >= 5:
            absences.append((expected, int(recent_count), species))

    for expected, recent_count, species in sorted(absences, reverse=True)[:3]:
        add_insight(
            insights,
            71,
            "Absence",
            f"No {species} detections in this period",
            f"{recent_count:,} detections in the previous {lookback_days} days, but none in the selected period.",
            species,
            chart={"type": "species_recent", "species": species},
        )


def add_community_mix_insights(all_news, current_news, start_date, end_date, insights):
    if len(current_news) == 0:
        return

    total = len(current_news)
    species_counts = current_news["Com_Name"].value_counts()
    if len(species_counts) == 0:
        return

    top_species = species_counts.index[0]
    top_count = int(species_counts.iloc[0])
    top_share = top_count / total
    if total >= 20 and top_share >= 0.50:
        add_insight(
            insights,
            74,
            "Community mix",
            f"{top_species} dominated the soundscape",
            f"{top_count:,} of {total:,} detections ({top_share:.0%}) were {top_species}.",
            top_species,
            chart={"type": "species_recent", "species": top_species},
        )

    if "Diet" in current_news.columns:
        diet_counts = current_news[current_news["Diet"] != "Unclassified"]["Diet"].value_counts()
        if len(diet_counts):
            top_diet = diet_counts.index[0]
            diet_share = diet_counts.iloc[0] / max(diet_counts.sum(), 1)
            if diet_counts.iloc[0] >= 20 and diet_share >= 0.60:
                add_insight(
                    insights,
                    60,
                    "Community mix",
                    f"{top_diet}s dominated detections",
                    f"{diet_share:.0%} of classified detections were {str(top_diet).lower()} species.",
                )

    historical = all_news[
        ~((all_news["date"] >= start_date) & (all_news["date"] <= end_date))
    ].copy()
    if total >= 25 and len(historical):
        current_season = news_season_from_month(end_date.month)
        current_counts = current_news["Com_Name"].value_counts()
        season_scores = []
        for season in ["Spring", "Summer", "Autumn", "Winter"]:
            season_data = historical[historical["month_num"].apply(news_season_from_month) == season]
            if season_data["date"].nunique() < 7:
                continue
            score = species_mix_similarity(current_counts, season_data["Com_Name"].value_counts())
            season_scores.append((score, season))

        if len(season_scores) >= 2:
            season_scores = sorted(season_scores, reverse=True)
            best_score, best_season = season_scores[0]
            current_score = next((score for score, season in season_scores if season == current_season), 0)
            if best_season != current_season and best_score >= 0.55 and best_score >= current_score + 0.12:
                add_insight(
                    insights,
                    59,
                    "Community mix",
                    f"Species mix looked more like {best_season.lower()}",
                    f"The current species mix matched historic {best_season.lower()} patterns more closely than {current_season.lower()}.",
                )

    period_days = (end_date - start_date).days + 1
    all_dates = pd.date_range(all_news["date"].min(), all_news["date"].max(), freq="D").date
    species_by_date = all_news.groupby("date")["Com_Name"].apply(lambda x: set(x.dropna())).to_dict()
    rolling_species = []
    for idx, day in enumerate(all_dates):
        if idx + 1 < period_days:
            continue
        window_species = set()
        for window_day in all_dates[idx + 1 - period_days:idx + 1]:
            window_species.update(species_by_date.get(window_day, set()))
        rolling_species.append({"date": day, "species_count": len(window_species)})

    rolling_df = pd.DataFrame(rolling_species)
    if len(rolling_df) >= 5:
        current_species = current_news["Com_Name"].nunique()
        before_current = rolling_df[rolling_df["date"] < end_date]
        median_species = rolling_df["species_count"].median()
        if len(before_current) and current_species >= before_current["species_count"].max() and current_species > 0:
            add_insight(
                insights,
                83,
                "Community mix",
                f"Richest species mix for a {period_noun(period_days)}",
                f"{current_species:,} species detected in the selected period.",
                chart={"type": "species_mix_period"},
            )
        elif median_species > 0:
            change = pct_change(current_species, median_species)
            if change is not None and change <= -35:
                add_insight(
                    insights,
                    65,
                    "Community mix",
                    "Species mix was narrower than usual",
                    f"{current_species:,} species, {signed_pct_label(change)} vs the typical {period_noun(period_days)}.",
                    chart={"type": "species_mix_period"},
                )


def add_time_of_day_insights(all_news, current_news, start_date, end_date, insights):
    if len(current_news) == 0:
        return

    period_days = (end_date - start_date).days + 1
    total = len(current_news)
    early_count = len(current_news[current_news["hour"] < 7])
    if total >= 20 and early_count / total >= 0.50:
        add_insight(
            insights,
            68,
            "Time of day",
            "Most activity happened before 7am",
            f"{early_count:,} of {total:,} detections ({early_count / total:.0%}) were before 07:00.",
            chart={"type": "hourly_activity", "highlight_hours": list(range(0, 7))},
        )

    hour_counts = current_news["hour"].value_counts()
    if len(hour_counts):
        peak_hour = int(hour_counts.index[0])
        peak_count = int(hour_counts.iloc[0])
        peak_share = peak_count / total
        if total >= 25 and peak_share >= 0.35:
            add_insight(
                insights,
                69,
                "Time of day",
                "Activity concentrated into one noisy hour",
                f"{peak_count:,} detections ({peak_share:.0%}) came during {peak_hour:02d}:00-{(peak_hour + 1) % 24:02d}:00.",
                chart={"type": "hourly_activity", "highlight_hours": [peak_hour]},
            )

    comp = comparison_window(all_news, current_news, start_date, end_date)
    comp_days = comp["date"].nunique()
    if comp_days == 0:
        return

    for label, mask_fn, priority, highlight_hours in [
        ("evening", lambda data: (data["hour"] >= 17) & (data["hour"] < 22), 66, list(range(17, 22))),
        ("night", lambda data: (data["hour"] >= 22) | (data["hour"] < 5), 64, [22, 23, 0, 1, 2, 3, 4]),
    ]:
        current_count = int(mask_fn(current_news).sum())
        expected = (int(mask_fn(comp).sum()) / comp_days) * period_days
        if expected >= 5 and current_count >= max(expected * 2.0, expected + 10):
            add_insight(
                insights,
                priority,
                "Time of day",
                f"Unusual {label} activity",
                f"{current_count:,} {label} detections vs about {expected:.0f} expected.",
                chart={"type": "hourly_activity", "highlight_hours": highlight_hours},
            )


def add_weather_insights(all_news, current_news, start_date, end_date, weather_daily, insights):
    if weather_daily is None or len(weather_daily) == 0 or len(current_news) == 0:
        return

    period_days = (end_date - start_date).days + 1
    typical = expected_count_for_period(all_news, current_news, start_date, end_date)
    current_total = len(current_news)
    rain_total = weather_daily["precip_sum"].sum()
    peak_wind = weather_daily["wind_max"].max()

    if typical is not None and typical > 0:
        activity_change = pct_change(current_total, typical)
        if rain_total >= max(5.0, period_days * 1.5) and current_total >= typical * 1.25:
            add_insight(
                insights,
                67,
                "Weather",
                "Rain did not dampen activity",
                f"{rain_total:.1f} mm of rain, but detections were {signed_pct_label(activity_change)} vs expected.",
                chart={"type": "weather_activity", "weather_metric": "precip_sum"},
            )
        if peak_wind >= 35 and current_total <= typical * 0.75:
            add_insight(
                insights,
                67,
                "Weather",
                "Strong wind coincided with quieter activity",
                f"Peak wind was {peak_wind:.1f} km/h and detections were {signed_pct_label(activity_change)} vs expected.",
                chart={"type": "weather_activity", "weather_metric": "wind_max"},
            )

    if period_days >= 3:
        daily_activity = (
            current_news.groupby("date")
            .agg(det_count=("Com_Name", "size"), species_count=("Com_Name", "nunique"))
            .reset_index()
        )
        merged = daily_activity.merge(weather_daily, on="date", how="inner")
        if len(merged) >= 3:
            warmest = merged.loc[merged["temp_max"].idxmax()]
            if warmest["species_count"] == merged["species_count"].max() and warmest["species_count"] >= merged["species_count"].median() + 2:
                add_insight(
                    insights,
                    61,
                    "Weather",
                    "Warmest day had the richest species mix",
                    f"{date_label(warmest['date'])}: {warmest['temp_max']:.1f}C max and {int(warmest['species_count'])} species.",
                    chart={"type": "weather_activity", "weather_metric": "temp_max"},
                )
            wettest = merged.loc[merged["precip_sum"].idxmax()]
            if wettest["precip_sum"] >= 3 and wettest["det_count"] <= merged["det_count"].median() * 0.6:
                add_insight(
                    insights,
                    58,
                    "Weather",
                    "Wettest day was one of the quietest",
                    f"{date_label(wettest['date'])}: {wettest['precip_sum']:.1f} mm rain and {int(wettest['det_count'])} detections.",
                    chart={"type": "weather_activity", "weather_metric": "precip_sum"},
                )


def add_record_streak_insights(all_news, current_news, start_date, end_date, insights):
    if len(all_news) == 0 or len(current_news) == 0:
        return

    current_species = set(current_news["Com_Name"].dropna().unique())
    streaks = []
    for species in current_species:
        species_dates = all_news[all_news["Com_Name"] == species]["date"].tolist()
        current_streak = streak_ending_at(species_dates, end_date)
        best_streak = longest_date_streak(species_dates)
        if current_streak >= 7 and current_streak >= best_streak:
            streaks.append((current_streak, species, True))
        elif current_streak >= 14:
            streaks.append((current_streak, species, False))

    for current_streak, species, is_record in sorted(streaks, reverse=True)[:3]:
        headline = f"Longest detection streak for {species}" if is_record else f"{species} streak continues"
        detail = f"Detected on {current_streak} consecutive days up to {date_label(end_date)}."
        add_insight(
            insights,
            79 if is_record else 66,
            "Record",
            headline,
            detail,
            species,
            chart={"type": "species_recent", "species": species},
        )

    if "Confidence" in current_news.columns and current_news["Confidence"].notna().any():
        period_days = (end_date - start_date).days + 1
        comp = comparison_window(all_news.dropna(subset=["Confidence"]), current_news, start_date, end_date)
        if len(comp) and comp["Confidence"].notna().any():
            current_conf = current_news["Confidence"].mean()
            comp_conf = comp["Confidence"].mean()
            if current_conf >= 0.85 and current_conf >= comp_conf + 0.08 and len(current_news) >= 10:
                add_insight(
                    insights,
                    57,
                    "Record",
                    "High-confidence detection period",
                    f"Average confidence was {current_conf:.2f}, above the comparison average of {comp_conf:.2f}.",
                )

        if period_days == 1:
            current_species_set = set(current_news["Com_Name"].dropna().astype(str).value_counts().head(3).index)
            if len(current_species_set) >= 3:
                previous_sets = daily_species_sets(all_news[all_news["date"] < start_date])
                seen_before = any(current_species_set.issubset(species_set) for species_set in previous_sets.values())
                if not seen_before:
                    add_insight(
                        insights,
                        55,
                        "Record",
                        "New top-three species combination",
                        "The three most-recorded species for this day had not previously been detected together on a single day.",
                    )


def add_data_quality_insights(all_news, current_news, start_date, end_date, insights):
    if len(current_news) == 0:
        return

    if "UK_Status" in current_news.columns:
        review_statuses = ["Review Recording", "False Positive"]
        review_rows = current_news[current_news["UK_Status"].isin(review_statuses)]
        if len(review_rows) >= 5:
            add_insight(
                insights,
                86,
                "Data quality",
                "Several detections need review",
                f"{len(review_rows):,} detections are marked as Review Recording or False Positive in this period.",
            )

        rare_statuses = ["Rare vagrant", "Scarce visitor"]
        rare_rows = current_news[current_news["UK_Status"].isin(rare_statuses)].copy()
        if len(rare_rows):
            low_conf_rare = rare_rows[rare_rows["Confidence"] <= 0.75] if "Confidence" in rare_rows.columns else rare_rows.iloc[0:0]
            top_rare = rare_rows["Com_Name"].value_counts().head(3).index.tolist()
            if len(low_conf_rare) or len(rare_rows) >= 3:
                add_insight(
                    insights,
                    82,
                    "Data quality",
                    "Unusual detections may need review",
                    f"{len(rare_rows):,} scarce or rare detections: {', '.join(top_rare)}.",
                )

    if "Confidence" in current_news.columns and current_news["Confidence"].notna().any():
        period_days = (end_date - start_date).days + 1
        low_conf = current_news[current_news["Confidence"] <= 0.70]
        low_rate = len(low_conf) / max(len(current_news), 1)
        comp = comparison_window(all_news.dropna(subset=["Confidence"]), current_news, start_date, end_date)
        comp_low_rate = (comp["Confidence"] <= 0.70).mean() if len(comp) else 0
        if len(low_conf) >= 10 and low_rate >= max(0.20, comp_low_rate * 2):
            add_insight(
                insights,
                78,
                "Data quality",
                "Low-confidence detections spiked",
                f"{len(low_conf):,} low-confidence detections ({low_rate:.0%}) vs a comparison rate of {comp_low_rate:.0%}.",
            )


def add_garden_event_watch_insights(all_news, current_news, end_date, events, insights):
    if len(all_news) == 0:
        return

    year = end_date.year
    month = end_date.month
    year_df = all_news[
        (all_news["year"] == year) &
        (all_news["date"] <= end_date)
    ].copy()
    for event in events:
        event_months = set(event.get("months", []))
        if event_months and month not in event_months:
            continue

        if event.get("trigger") == "first_seen_year":
            seen_this_year = int(event_species_mask(year_df, event).sum()) > 0
            if not seen_this_year:
                add_insight(
                    insights,
                    52,
                    "Garden year",
                    f"{event.get('label', 'Arrival window')} window is open",
                    "No matching detections yet this year, but this is the usual seasonal window.",
                )
        elif event.get("trigger") == "spike" and len(current_news):
            current_count = int(event_species_mask(current_news, event).sum())
            if current_count >= 5:
                add_insight(
                    insights,
                    50,
                    "Garden year",
                    f"{event.get('label', 'Garden event')} is active",
                    f"{current_count:,} matching detections in the selected period.",
                )


def build_news_insights(all_data, period_data, start_date, end_date, events, weather_daily=None):
    all_news = prepare_news_df(all_data)
    current_news = prepare_news_df(period_data)
    insights = []

    if start_date is None or end_date is None:
        return insights
    if len(current_news) == 0:
        return insights

    add_period_record_insights(all_news, current_news, start_date, end_date, insights)
    add_arrival_insights(all_news, current_news, start_date, end_date, events, insights)
    add_expected_arrival_insights(all_news, current_news, start_date, end_date, insights)
    add_species_change_insights(all_news, current_news, start_date, end_date, events, insights)
    add_event_ytd_insights(all_news, end_date, events, insights)
    add_dawn_chorus_insights(all_news, current_news, start_date, end_date, weather_daily, insights)
    add_absence_comeback_insights(all_news, current_news, start_date, end_date, insights)
    add_community_mix_insights(all_news, current_news, start_date, end_date, insights)
    add_time_of_day_insights(all_news, current_news, start_date, end_date, insights)
    add_weather_insights(all_news, current_news, start_date, end_date, weather_daily, insights)
    add_record_streak_insights(all_news, current_news, start_date, end_date, insights)
    add_data_quality_insights(all_news, current_news, start_date, end_date, insights)
    add_garden_event_watch_insights(all_news, current_news, end_date, events, insights)

    seen = set()
    deduped = []
    for insight in sorted(insights, key=lambda x: x["priority"], reverse=True):
        key = insight_key(insight)
        if key in seen:
            continue
        seen.add(key)
        deduped.append(insight)
    return deduped


def select_visible_news_insights(news_insights, limit=VISIBLE_HEADLINE_LIMIT):
    if len(news_insights) <= limit:
        return news_insights

    selected = news_insights[:limit]

    for candidate in news_insights[limit:]:
        category = candidate["category"]
        visible_categories = {insight["category"] for insight in selected}
        if category in visible_categories:
            continue

        category_counts = pd.Series([insight["category"] for insight in selected]).value_counts().to_dict()
        replace_idx = None
        for idx in range(len(selected) - 1, -1, -1):
            selected_category = selected[idx]["category"]
            if category_counts.get(selected_category, 0) > 1:
                replace_idx = idx
                break

        if replace_idx is None:
            selected.append(candidate)
        else:
            selected[replace_idx] = candidate

    selected = sorted(selected, key=lambda insight: news_insights.index(insight))
    return selected


def news_chart_start(all_news, end_date, period_days):
    lookback_days = max(60, period_days * 4)
    return max(all_news["date"].min(), end_date - datetime.timedelta(days=lookback_days))


def style_news_fig(fig):
    fig = style_fig(fig)
    fig.update_layout(
        height=NEWS_CHART_HEIGHT,
        margin=dict(l=8, r=8, t=28, b=6),
        title_text="",
        legend_title_text="",
        legend=dict(
            orientation="h",
            y=1.08,
            yanchor="bottom",
            x=0.01,
            font=dict(size=11, color="#1a2416"),
            title=dict(text=""),
        ),
    )
    fig.update_xaxes(title_font=dict(size=11), tickfont=dict(size=10))
    fig.update_yaxes(title_font=dict(size=11), tickfont=dict(size=10))
    return fig


def render_activity_period_chart(all_data, start_date, end_date, chart_key):
    all_news = prepare_news_df(all_data)
    if len(all_news) == 0:
        return False

    period_days = (end_date - start_date).days + 1
    all_dates = pd.date_range(all_news["date"].min(), end_date, freq="D").date
    daily_counts = all_news.groupby("date").size().reindex(all_dates, fill_value=0)

    if period_days == 1:
        series = daily_counts
        y_title = "Daily detections"
    else:
        series = daily_counts.rolling(period_days, min_periods=period_days).sum().dropna()
        y_title = f"{period_days}-day detections"

    chart_start = news_chart_start(all_news, end_date, period_days)
    chart_df = series.rename("Detections").reset_index().rename(columns={"index": "date"})
    chart_df = chart_df[chart_df["date"] >= chart_start].copy()
    if len(chart_df) == 0:
        return False

    chart_df["Selected"] = chart_df["date"].apply(
        lambda d: "Selected" if start_date <= d <= end_date else "Other"
    )
    if period_days > 1:
        chart_df["Selected"] = chart_df["date"].apply(lambda d: "Selected" if d == end_date else "Other")

    fig = px.bar(
        chart_df,
        x="date",
        y="Detections",
        color="Selected",
        title="",
        labels={"date": "Date", "Detections": y_title, "Selected": ""},
        color_discrete_map={"Selected": TERTIARY, "Other": PRIMARY},
    )
    median_value = float(series.median()) if len(series) else 0
    fig.add_scatter(
        x=[chart_df["date"].min(), chart_df["date"].max()],
        y=[median_value, median_value],
        mode="lines",
        line=dict(color="#1a2416", width=2, dash="dot"),
        name="Median",
    )
    fig.update_layout(showlegend=True)
    fig.update_traces(marker_line_width=0)
    st.plotly_chart(style_news_fig(fig), width="stretch", config={"displayModeBar": False}, key=chart_key)
    return True


def render_species_recent_chart(all_data, start_date, end_date, species, chart_key):
    all_news = prepare_news_df(all_data)
    if len(all_news) == 0 or not species:
        return False

    species_news = all_news[all_news["Com_Name"] == species].copy()
    if len(species_news) == 0:
        return False

    period_days = (end_date - start_date).days + 1
    chart_start = news_chart_start(all_news, end_date, period_days)
    chart_dates = pd.date_range(chart_start, end_date, freq="D").date
    species_counts = species_news.groupby("date").size().reindex(chart_dates, fill_value=0)
    chart_df = species_counts.rename("Detections").reset_index().rename(columns={"index": "date"})
    chart_df["Selected"] = chart_df["date"].apply(
        lambda d: "Selected" if start_date <= d <= end_date else "Other"
    )

    comparison = comparison_window(all_news, prepare_news_df(filter_date_window(all_data, start_date, end_date)), start_date, end_date)
    expected_daily = None
    comp_species = comparison[comparison["Com_Name"] == species].copy()
    comp_days = comparison["date"].nunique()
    if comp_days:
        expected_daily = len(comp_species) / comp_days

    fig = px.bar(
        chart_df,
        x="date",
        y="Detections",
        color="Selected",
        title="",
        labels={"date": "Date", "Detections": "Detections", "Selected": ""},
        color_discrete_map={"Selected": TERTIARY, "Other": SECONDARY},
    )
    if expected_daily is not None:
        fig.add_scatter(
            x=[chart_df["date"].min(), chart_df["date"].max()],
            y=[expected_daily, expected_daily],
            mode="lines",
            line=dict(color="#1a2416", width=2, dash="dot"),
            name="Expected/day",
        )
    fig.update_layout(showlegend=True)
    fig.update_traces(marker_line_width=0)
    st.plotly_chart(style_news_fig(fig), width="stretch", config={"displayModeBar": False}, key=chart_key)
    return True


def render_species_mix_period_chart(all_data, start_date, end_date, chart_key):
    all_news = prepare_news_df(all_data)
    if len(all_news) == 0:
        return False

    period_days = (end_date - start_date).days + 1
    all_dates = pd.date_range(all_news["date"].min(), end_date, freq="D").date
    species_by_date = all_news.groupby("date")["Com_Name"].apply(lambda x: set(x.dropna())).to_dict()
    rolling_species = []
    for idx, day in enumerate(all_dates):
        if idx + 1 < period_days:
            continue
        window_species = set()
        for window_day in all_dates[idx + 1 - period_days:idx + 1]:
            window_species.update(species_by_date.get(window_day, set()))
        rolling_species.append({"date": day, "Species": len(window_species)})

    chart_df = pd.DataFrame(rolling_species)
    if len(chart_df) == 0:
        return False

    chart_start = news_chart_start(all_news, end_date, period_days)
    chart_df = chart_df[chart_df["date"] >= chart_start].copy()
    if len(chart_df) == 0:
        return False

    chart_df["Selected"] = chart_df["date"].apply(
        lambda d: "Selected" if start_date <= d <= end_date else "Other"
    )
    if period_days > 1:
        chart_df["Selected"] = chart_df["date"].apply(lambda d: "Selected" if d == end_date else "Other")

    y_title = "Daily species" if period_days == 1 else f"{period_days}-day species"
    fig = px.bar(
        chart_df,
        x="date",
        y="Species",
        color="Selected",
        title="",
        labels={"date": "Date", "Species": y_title, "Selected": ""},
        color_discrete_map={"Selected": TERTIARY, "Other": SECONDARY},
    )
    median_value = float(chart_df["Species"].median()) if len(chart_df) else 0
    fig.add_scatter(
        x=[chart_df["date"].min(), chart_df["date"].max()],
        y=[median_value, median_value],
        mode="lines",
        line=dict(color="#1a2416", width=2, dash="dot"),
        name="Median",
    )
    fig.update_layout(showlegend=True)
    fig.update_traces(marker_line_width=0)
    st.plotly_chart(style_news_fig(fig), width="stretch", config={"displayModeBar": False}, key=chart_key)
    return True


def render_hourly_activity_chart(all_data, period_data, start_date, end_date, chart_key, highlight_hours=None):
    all_news = prepare_news_df(all_data)
    current_news = prepare_news_df(period_data)
    if len(all_news) == 0 or len(current_news) == 0:
        return False

    period_days = (end_date - start_date).days + 1
    hours = list(range(24))
    current_counts = current_news.groupby("hour").size().reindex(hours, fill_value=0)
    highlight_hours = set(highlight_hours or [])
    bar_colors = [TERTIARY if hour in highlight_hours else PRIMARY for hour in hours]
    hour_labels = [f"{hour:02d}:00" for hour in hours]

    fig = go.Figure()
    fig.add_trace(
        go.Bar(
            x=hour_labels,
            y=current_counts.values,
            name="Selected period",
            marker_color=bar_colors,
            opacity=0.78,
        )
    )

    comp = comparison_window(all_news, current_news, start_date, end_date)
    comp_days = comp["date"].nunique()
    if comp_days:
        expected_counts = comp.groupby("hour").size().reindex(hours, fill_value=0) / comp_days * period_days
        fig.add_trace(
            go.Scatter(
                x=hour_labels,
                y=expected_counts.values,
                name="Expected",
                mode="lines+markers",
                line=dict(color="#1a2416", width=2.5),
                marker=dict(size=5, color="#1a2416"),
            )
        )

    fig.update_layout(
        title_text="",
        showlegend=True,
        xaxis_title="Hour",
        yaxis_title="Detections",
    )
    st.plotly_chart(style_news_fig(fig), width="stretch", config={"displayModeBar": False}, key=chart_key)
    return True


def render_weather_activity_chart(period_data, weather_daily, weather_metric, chart_key):
    if weather_daily is None or len(weather_daily) == 0:
        return False

    current_news = prepare_news_df(period_data)
    if len(current_news) == 0:
        return False

    daily_activity = (
        current_news.groupby("date")
        .agg(detections=("Com_Name", "size"), species=("Com_Name", "nunique"))
        .reset_index()
    )
    merged = daily_activity.merge(weather_daily, on="date", how="inner")
    if len(merged) == 0 or weather_metric not in merged.columns:
        return False

    metric_labels = {
        "precip_sum": ("Rainfall", "Rainfall (mm)", SECONDARY),
        "wind_max": ("Max wind", "Max wind (km/h)", "#607080"),
        "temp_max": ("Max temp", "Max temp (C)", "#c47a5a"),
    }
    metric_name, metric_label, metric_color = metric_labels.get(
        weather_metric,
        (weather_metric, weather_metric, SECONDARY),
    )

    fig = make_subplots(specs=[[{"secondary_y": True}]])
    fig.add_trace(
        go.Bar(
            x=merged["date"],
            y=merged["detections"],
            name="Detections",
            marker_color=PRIMARY,
            opacity=0.72,
        ),
        secondary_y=False,
    )
    fig.add_trace(
        go.Scatter(
            x=merged["date"],
            y=merged[weather_metric],
            name=metric_name,
            mode="lines+markers",
            line=dict(color=metric_color, width=2.5),
            marker=dict(size=7, color=metric_color),
        ),
        secondary_y=True,
    )
    fig.update_layout(title_text="", legend=dict(x=0.01, y=0.99))
    fig.update_yaxes(title_text="Detections", secondary_y=False)
    fig.update_yaxes(title_text=metric_label, secondary_y=True)
    st.plotly_chart(style_news_fig(fig), width="stretch", config={"displayModeBar": False}, key=chart_key)
    return True


def headline_chart_key(insight, start_date, end_date, index):
    payload = {
        "category": insight.get("category"),
        "headline": insight.get("headline"),
        "species": insight.get("species"),
        "chart": insight.get("chart"),
        "start_date": str(start_date),
        "end_date": str(end_date),
        "index": index,
    }
    digest = hashlib.md5(json.dumps(payload, sort_keys=True, default=str).encode("utf-8")).hexdigest()
    return f"headline_chart_{digest}"


def render_news_chart(insight, all_data, period_data, start_date, end_date, weather_daily, chart_key):
    chart = insight.get("chart")
    if not chart:
        return False

    chart_type = chart.get("type")
    if chart_type == "activity_period":
        return render_activity_period_chart(all_data, start_date, end_date, chart_key)
    if chart_type == "species_recent":
        return render_species_recent_chart(all_data, start_date, end_date, chart.get("species"), chart_key)
    if chart_type == "species_mix_period":
        return render_species_mix_period_chart(all_data, start_date, end_date, chart_key)
    if chart_type == "hourly_activity":
        return render_hourly_activity_chart(
            all_data,
            period_data,
            start_date,
            end_date,
            chart_key,
            chart.get("highlight_hours"),
        )
    if chart_type == "weather_activity":
        return render_weather_activity_chart(period_data, weather_daily, chart.get("weather_metric"), chart_key)
    return False


def render_news_insight(insight, all_data=None, period_data=None, start_date=None, end_date=None, weather_daily=None, index=0):
    headline = html.escape(str(insight["headline"]))
    detail = html.escape(str(insight["detail"]))
    accent = NEWS_CATEGORY_COLORS.get(insight["category"], PRIMARY)
    st.markdown(
        f"""
<div class="news-card" style="--news-accent: {accent};">
  <div class="news-headline">{headline}</div>
  <div class="news-detail">{detail}</div>
</div>
""",
        unsafe_allow_html=True,
    )
    if insight.get("chart") and all_data is not None and period_data is not None:
        chart_key = headline_chart_key(insight, start_date, end_date, index)
        render_news_chart(insight, all_data, period_data, start_date, end_date, weather_daily, chart_key)


st.title("🐦 Garden Bird Dashboard")
st.caption("Detections across time, seasons, and community composition.")

# ---- Sidebar filters ----
st.sidebar.header("Explore")

_pages = [
    "Daily Overview",
    "Overview",
    "Community",
    "NMDS",
    "Dawn Chorus Overview",
    "Weather & Activity",
    "Data Quality",
    "Records",
    "Nearby Sightings",
    "Species Explorer",
]
if datetime.date.today().month == 2 and datetime.date.today().day == 23:
    _pages.append("\U0001f382")

page = st.sidebar.radio(
    "View",
    _pages,
    label_visibility="collapsed",
)
st.sidebar.divider()

min_conf = st.sidebar.slider(
    "Minimum Confidence",
    float(df["Confidence"].min()),
    float(df["Confidence"].max()),
    float(df["Confidence"].min()),
)

filtered = df[df["Confidence"] >= min_conf].copy()

species_list = st.sidebar.multiselect(
    "Select Species",
    sorted(filtered["Com_Name"].dropna().unique()),
)
if species_list:
    filtered = filtered[filtered["Com_Name"].isin(species_list)]

status_list = st.sidebar.multiselect(
    "UK Status",
    sorted(filtered["UK_Status"].dropna().unique()),
)
if status_list:
    filtered = filtered[filtered["UK_Status"].isin(status_list)]

filtered_pre_date = filtered.copy()

st.sidebar.subheader("Date Range")
filtered_ts = filtered.dropna(subset=["timestamp"]).copy()
min_date = filtered_ts["timestamp"].min().date()
max_date = filtered_ts["timestamp"].max().date()

date_range = st.sidebar.date_input(
    "Select Date Range",
    (min_date, max_date),
    min_value=min_date,
    max_value=max_date,
)

if isinstance(date_range, tuple) and len(date_range) == 2:
    start_date, end_date = date_range
else:
    start_date = end_date = date_range

filtered = filtered[
    (filtered["timestamp"].dt.date >= start_date) &
    (filtered["timestamp"].dt.date <= end_date)
].copy()

def season_from_month(m: int) -> str:
    if m in (3, 4, 5):  return "Spring"
    if m in (6, 7, 8):  return "Summer"
    if m in (9, 10, 11): return "Autumn"
    return "Winter"

filtered["year"]      = filtered["timestamp"].dt.year.astype("Int64")
filtered["month_num"] = filtered["timestamp"].dt.month.astype("Int64")
filtered["season"]    = filtered["month_num"].apply(lambda m: season_from_month(m) if pd.notna(m) else None)

# ── Year / Season / Month sidebar filters ──
st.sidebar.subheader("Year / Season / Month")

_years_available = sorted(filtered["year"].dropna().unique())
year_mode = st.sidebar.selectbox("Years", ["All years", "Select years"], index=0)
selected_years = []
if year_mode == "Select years":
    selected_years = st.sidebar.multiselect(
        "Choose years", _years_available,
        default=_years_available[-1:] if _years_available else [],
    )
    if selected_years:
        filtered = filtered[filtered["year"].isin(selected_years)].copy()

selected_season = st.sidebar.selectbox("Season", ["All", "Spring", "Summer", "Autumn", "Winter"], index=0)

MONTHS_FULL = [
    (1, "January"), (2, "February"), (3, "March"), (4, "April"),
    (5, "May"), (6, "June"), (7, "July"), (8, "August"),
    (9, "September"), (10, "October"), (11, "November"), (12, "December"),
]
month_num_by_name = {n: m for m, n in MONTHS_FULL}
month_names_list  = [name for _, name in MONTHS_FULL]

month_mode = st.sidebar.selectbox("Months", ["All months", "Choose month"], index=0)
chosen_month = None
if month_mode == "Choose month":
    chosen_month = st.sidebar.selectbox("Choose month value", month_names_list, index=0)

# Keep a copy before season/month filters for compare-mode overrides
_filtered_pre_season_month = filtered.copy()

if selected_season != "All":
    filtered = filtered[filtered["season"] == selected_season].copy()
if month_mode == "Choose month" and chosen_month:
    filtered = filtered[filtered["month_num"] == month_num_by_name[chosen_month]].copy()

st.sidebar.divider()

exclude_review = st.sidebar.checkbox("Exclude 'Review Recording' & 'False Positive'", value=True)
review_df = filtered[filtered["UK_Status"] == "Review Recording"].copy()
if exclude_review:
    filtered = filtered[~filtered["UK_Status"].isin(["Review Recording", "False Positive"])].copy()

daily_base = filtered_pre_date.copy()
if exclude_review:
    daily_base = daily_base[~daily_base["UK_Status"].isin(["Review Recording", "False Positive"])].copy()

daily_base = daily_base.dropna(subset=["timestamp"]).copy()
daily_available_dates = sorted(daily_base["timestamp"].dt.date.unique().tolist())

def default_daily_overview_date(available_dates):
    if not available_dates:
        return None
    yesterday = datetime.date.today() - datetime.timedelta(days=1)
    eligible_dates = [d for d in available_dates if d <= yesterday]
    return eligible_dates[-1] if eligible_dates else available_dates[-1]

daily_selected_date = None
if daily_available_dates:
    _daily_default = default_daily_overview_date(daily_available_dates)
    if st.session_state.get("daily_overview_date") not in daily_available_dates:
        st.session_state["daily_overview_date"] = _daily_default
    daily_selected_date = st.session_state["daily_overview_date"]

daily_period_mode = st.session_state.get("daily_period_mode", DAILY_PERIOD_OPTIONS[0])
if daily_period_mode not in DAILY_PERIOD_OPTIONS:
    daily_period_mode = DAILY_PERIOD_OPTIONS[0]
daily_window_start, daily_window_end = daily_period_bounds(daily_selected_date, daily_period_mode)
daily_period_filtered = filter_date_window(daily_base, daily_window_start, daily_window_end)

# ---- KPI cards ----
kpi_source = daily_period_filtered if page == "Daily Overview" else filtered
kpi1, kpi2, kpi3 = st.columns(3)
kpi1.metric("Total Detections",  f"{len(kpi_source):,}")
kpi2.metric("Unique Species",    f"{kpi_source['Com_Name'].nunique():,}")
kpi3.metric("Average Confidence",
            f"{kpi_source['Confidence'].mean():.2f}" if len(kpi_source) else "—")

st.divider()

# ── Daily Overview ──────────────────────────────────────────────────────────
if page == "Daily Overview":
    st.markdown('<div class="news-feed-title">Garden News Feed</div>', unsafe_allow_html=True)

    if not daily_available_dates:
        st.info("No dated detections are available for the current species, status, and confidence filters.")
    else:
        current_idx = daily_available_dates.index(daily_selected_date)
        date_col, controls_col = st.columns([1.7, 1], gap="large")
        with controls_col:
            daily_period_mode = st.radio(
                "Period",
                DAILY_PERIOD_OPTIONS,
                horizontal=True,
                key="daily_period_mode",
                label_visibility="collapsed",
            )

        daily_window_start, daily_window_end = daily_period_bounds(daily_selected_date, daily_period_mode)
        daily_period_filtered = filter_date_window(daily_base, daily_window_start, daily_window_end)

        with date_col:
            st.markdown(
                f"""
<div class="daily-date-panel">
  <div class="daily-date-title">{html.escape(format_period_label(daily_window_start, daily_window_end))}</div>
</div>
""",
                unsafe_allow_html=True,
            )
        with controls_col:
            st.markdown('<div class="daily-controls-spacer"></div>', unsafe_allow_html=True)
            prev_col, latest_col, next_col = st.columns([0.55, 1.2, 0.55], gap="small")
            with prev_col:
                if st.button("‹", use_container_width=True, disabled=current_idx == 0, help="Previous day", key="daily_prev_day"):
                    st.session_state["daily_overview_date"] = daily_available_dates[current_idx - 1]
                    st.rerun()
            with latest_col:
                if st.button("Latest", use_container_width=True, disabled=current_idx == len(daily_available_dates) - 1, key="daily_latest_day"):
                    st.session_state["daily_overview_date"] = daily_available_dates[-1]
                    st.rerun()
            with next_col:
                if st.button("›", use_container_width=True, disabled=current_idx == len(daily_available_dates) - 1, help="Next day", key="daily_next_day"):
                    st.session_state["daily_overview_date"] = daily_available_dates[current_idx + 1]
                    st.rerun()

        daily_view = daily_period_filtered.dropna(subset=["timestamp"]).copy()
        daily_view["hour"] = daily_view["timestamp"].dt.hour

        if len(daily_view) == 0:
            st.info("No detections were recorded for this period under the current filters.")
        else:
            weather_daily = None
            if {"Lat", "Lon"}.issubset(daily_view.columns) and daily_view["Lat"].notna().any() and daily_view["Lon"].notna().any():
                w_lat = float(daily_view["Lat"].mode().iloc[0])
                w_lon = float(daily_view["Lon"].mode().iloc[0])
                _, weather_daily = fetch_weather(
                    w_lat, w_lon,
                    daily_window_start.strftime("%Y-%m-%d"),
                    daily_window_end.strftime("%Y-%m-%d"),
                )

            garden_events = load_garden_events()
            news_insights = build_news_insights(
                daily_base,
                daily_view,
                daily_window_start,
                daily_window_end,
                garden_events,
                weather_daily,
            )

            visible_news_insights = select_visible_news_insights(news_insights)
            visible_news_keys = {insight_key(insight) for insight in visible_news_insights}
            if visible_news_insights:
                for insight_idx, insight in enumerate(visible_news_insights):
                    render_news_insight(
                        insight,
                        daily_base,
                        daily_view,
                        daily_window_start,
                        daily_window_end,
                        weather_daily,
                        index=insight_idx,
                    )
            else:
                st.info("No major changes detected for this period.")

            hidden_news_insights = [
                insight for insight in news_insights
                if insight_key(insight) not in visible_news_keys
            ]
            if hidden_news_insights:
                with st.expander("Explore more headlines"):
                    insight_rows = [
                        {
                            "Headline": insight["headline"],
                            "Detail": insight["detail"],
                        }
                        for insight in hidden_news_insights
                    ]
                    st.dataframe(pd.DataFrame(insight_rows), hide_index=True, use_container_width=True)

            st.divider()
            st.subheader("Drill-down")

            st.subheader("Weather Summary")
            if weather_daily is None or len(weather_daily) == 0:
                st.info("No weather summary is available for this period.")
            elif daily_window_start == daily_window_end:
                weather_match = weather_daily[weather_daily["date"] == daily_window_end]
                weather_daily_row = weather_match.iloc[0] if len(weather_match) else None
                if weather_daily_row is None:
                    st.info("No weather summary is available for this day.")
                else:
                    wx1, wx2, wx3 = st.columns(3)
                    wx1.metric("Max Temp", f"{weather_daily_row['temp_max']:.1f}°C")
                    wx2.metric("Min Temp", f"{weather_daily_row['temp_min']:.1f}°C")
                    wx3.metric("Rainfall", f"{weather_daily_row['precip_sum']:.1f} mm")
                    wx4, wx5 = st.columns(2)
                    wx4.metric("Max Wind", f"{weather_daily_row['wind_max']:.1f} km/h")
                    wx5.metric(
                        "Daylight",
                        (
                            f"{weather_daily_row['sunrise'].strftime('%H:%M')} to "
                            f"{weather_daily_row['sunset'].strftime('%H:%M')}"
                        ),
                    )
            else:
                wx1, wx2, wx3 = st.columns(3)
                wx1.metric("Avg Max Temp", f"{weather_daily['temp_max'].mean():.1f}°C")
                wx2.metric("Avg Min Temp", f"{weather_daily['temp_min'].mean():.1f}°C")
                wx3.metric("Total Rainfall", f"{weather_daily['precip_sum'].sum():.1f} mm")
                wx4, wx5 = st.columns(2)
                wx4.metric("Peak Wind", f"{weather_daily['wind_max'].max():.1f} km/h")
                wx5.metric("Weather Days", f"{len(weather_daily):,}")

            species_order = daily_view["Com_Name"].value_counts().index.tolist()
            species_color_map = {
                sp: NATURE_PALETTE[i % len(NATURE_PALETTE)]
                for i, sp in enumerate(species_order)
            }

            st.divider()

            species_counts = (
                daily_view["Com_Name"].value_counts()
                .rename_axis("Species")
                .reset_index(name="Count")
                .sort_values("Count", ascending=True)
            )
            fig = px.bar(
                species_counts, x="Count", y="Species", orientation="h",
                title="All Species by Detection Count",
                color="Count",
                color_continuous_scale=[[0, "#a3c47a"], [1, "#2d5233"]],
                labels={"Count": "Detections", "Species": ""},
            )
            fig.update_coloraxes(showscale=False)
            fig.update_traces(marker_line_width=0)
            fig.update_layout(height=max(420, len(species_counts) * 26))
            st.plotly_chart(style_fig(fig), width="stretch")

            sp_hour = (
                daily_view.groupby(["hour", "Com_Name"])
                .size()
                .reset_index(name="Count")
            )
            fig = px.area(
                sp_hour, x="hour", y="Count",
                color="Com_Name",
                title="Activity by Hour by Species",
                labels={"hour": "Hour of day", "Count": "Detections", "Com_Name": "Species"},
                category_orders={"Com_Name": species_order},
                color_discrete_map=species_color_map,
            )
            hourly_total = daily_view.groupby("hour").size().reset_index(name="Count")
            fig.add_scatter(
                x=hourly_total["hour"], y=hourly_total["Count"],
                mode="lines+markers",
                line=dict(color="#1a2416", width=2.5, dash="dot"),
                marker=dict(size=5, color="#1a2416"),
                name="Total", showlegend=True,
            )
            fig.update_layout(xaxis=dict(dtick=1))
            fig.update_traces(marker_line_width=0)
            st.plotly_chart(style_fig(fig), width="stretch")

            comp_hour = (
                daily_view.groupby(["hour", "Com_Name"])
                .size()
                .reset_index(name="Count")
            )
            comp_hour["Percent"] = (
                comp_hour.groupby("hour")["Count"]
                .transform(lambda x: (x / x.sum()) * 100)
            )
            fig = px.bar(
                comp_hour,
                x="hour", y="Percent",
                color="Com_Name",
                title="Community Composition by Hour (%)",
                labels={"hour": "Hour of day", "Percent": "% of detections", "Com_Name": "Species"},
                category_orders={"Com_Name": species_order},
                color_discrete_map=species_color_map,
            )
            fig.update_layout(barmode="stack", xaxis=dict(dtick=1))
            fig.update_traces(marker_line_width=0)
            st.plotly_chart(style_fig(fig), width="stretch")

            heatmap_counts = (
                daily_view.groupby(["Com_Name", "hour"])
                .size()
                .unstack(fill_value=0)
                .reindex(index=species_order, columns=range(24), fill_value=0)
            )
            fig = px.imshow(
                heatmap_counts.values,
                x=list(range(24)),
                y=heatmap_counts.index.tolist(),
                title="Species by Hour Heatmap",
                color_continuous_scale=HEATMAP_SCALE,
                labels={"x": "Hour of day", "y": "Species", "color": "Detections"},
                aspect="auto",
            )
            fig.update_layout(
                xaxis=dict(dtick=1),
                height=max(500, len(heatmap_counts.index) * 24),
                coloraxis_colorbar=dict(
                    title="Detections",
                    tickfont=dict(size=11, color="#4a5c44"),
                    title_font=dict(size=12, color="#4a5c44"),
                    thickness=14,
                ),
            )
            st.plotly_chart(style_fig(fig), width="stretch")

            species_windows = (
                daily_view.groupby("Com_Name")
                .agg(
                    First_Appearance=("timestamp", "min"),
                    Last_Appearance=("timestamp", "max"),
                    Detections=("Com_Name", "size"),
                    UK_Status=("UK_Status", lambda x: x.mode().iloc[0] if len(x.mode()) else "Review Recording"),
                )
                .reset_index()
                .rename(columns={"Com_Name": "Species"})
            )
            species_windows["Sort_Order"] = species_windows["Species"].map({sp: i for i, sp in enumerate(species_order)})
            species_windows = species_windows.sort_values("Sort_Order")

            same_moment = species_windows["First_Appearance"] == species_windows["Last_Appearance"]
            species_windows.loc[same_moment, "Last_Appearance"] = (
                species_windows.loc[same_moment, "Last_Appearance"] + pd.Timedelta(minutes=10)
            )
            species_windows["_start_str"] = species_windows["First_Appearance"].dt.strftime("%H:%M")
            species_windows["_end_str"] = species_windows["Last_Appearance"].dt.strftime("%H:%M")

            gantt_cmap = status_color_map(species_windows["UK_Status"].unique())
            fig = px.timeline(
                species_windows,
                x_start="First_Appearance",
                x_end="Last_Appearance",
                y="Species",
                color="UK_Status",
                color_discrete_map=gantt_cmap,
                hover_data={"Detections": True, "UK_Status": True},
                title=(
                    "First and Last Appearance by Species"
                    if daily_window_start == daily_window_end
                    else "First and Last Detection by Species"
                ),
                labels={"Species": "", "UK_Status": "UK Status"},
            )
            for trace in fig.data:
                mask = species_windows["UK_Status"].astype(str) == trace.name
                if mask.any():
                    trace.customdata = species_windows.loc[mask, ["_start_str", "_end_str", "Detections"]].values
                    trace.hovertemplate = (
                        "<b>%{y}</b><br>"
                        "First: %{customdata[0]}<br>"
                        "Last: %{customdata[1]}<br>"
                        "Detections: %{customdata[2]}"
                        "<extra>%{fullData.name}</extra>"
                    )
            timeline_start = pd.Timestamp.combine(daily_window_start, datetime.time(0, 0))
            timeline_end = pd.Timestamp.combine(daily_window_end, datetime.time(0, 0)) + pd.Timedelta(days=1)
            timeline_axis = (
                dict(
                    range=[timeline_start, timeline_end],
                    dtick=3600000,
                    tickformat="%H:%M",
                )
                if daily_window_start == daily_window_end
                else dict(
                    range=[timeline_start, timeline_end],
                    dtick=86400000,
                    tickformat="%d %b",
                )
            )
            fig.update_yaxes(
                categoryorder="array",
                categoryarray=list(reversed(species_windows["Species"].tolist())),
            )
            fig.update_layout(
                height=max(500, len(species_windows) * 24),
                xaxis_title="Time of day",
                yaxis_title="",
                xaxis=timeline_axis,
            )
            st.plotly_chart(style_fig(fig), width="stretch")

# ── Overview ────────────────────────────────────────────────────────────────
elif page == "Overview":
    top = (
        filtered["Com_Name"].value_counts()
        .head(20)
        .reset_index()
    )
    top.columns = ["Species", "Count"]
    top = top.sort_values("Count", ascending=True)  # most common at top of chart

    fig = px.bar(
        top, x="Count", y="Species", orientation="h",
        title="Top 20 Most Common Species",
        color="Count",
        color_continuous_scale=[[0, "#a3c47a"], [1, "#2d5233"]],
        labels={"Count": "Detections", "Species": ""},
    )
    fig.update_coloraxes(showscale=False)
    fig.update_traces(marker_line_width=0)
    st.plotly_chart(style_fig(fig), width="stretch")

    st.divider()

    # ── New Arrival Alerts ──
    st.subheader("New Arrival Alerts")

    na_df = filtered.dropna(subset=["timestamp"]).copy()
    na_df["year"] = na_df["timestamp"].dt.year

    available_years = sorted(na_df["year"].dropna().unique())
    if len(available_years) == 0:
        st.info("No data available for new arrival analysis.")
    else:
        na_years = st.multiselect(
            "Years to inspect", available_years,
            default=[available_years[-1]], key="na_years",
        )
        if not na_years:
            st.info("Select at least one year.")
        else:
            first_det = (
                na_df.groupby(["Com_Name", "year"])["timestamp"]
                .min()
                .reset_index(name="First_Seen")
            )
            first_det["First_Seen_Date"] = first_det["First_Seen"].dt.date

            # Compute the earliest year each species was ever seen
            ever_first = first_det.groupby("Com_Name")["year"].min().reset_index(name="First_Year_Ever")
            first_det = first_det.merge(ever_first, on="Com_Name", how="left")
            first_det["New_Species"] = first_det["year"] == first_det["First_Year_Ever"]

            display_df = first_det[first_det["year"].isin(na_years)].copy()
            display_df = display_df.sort_values("First_Seen_Date")

            cols = st.columns(len(na_years))
            for i, yr in enumerate(sorted(na_years)):
                yr_new = display_df[(display_df["year"] == yr) & (display_df["New_Species"])]
                cols[i].metric(f"New arrivals {yr}", len(yr_new))

            st.dataframe(
                display_df[["Com_Name", "year", "First_Seen_Date", "New_Species"]]
                .rename(columns={"Com_Name": "Species", "year": "Year",
                                 "First_Seen_Date": "First Seen", "New_Species": "New Species"}),
                hide_index=True,
            )

    st.divider()

    # ── Year List Progress ──
    st.subheader("Year List Progress")

    yl_df = filtered.dropna(subset=["timestamp"]).copy()
    yl_df["year"] = yl_df["timestamp"].dt.year
    yl_years_avail = sorted(yl_df["year"].dropna().unique())

    if len(yl_years_avail) < 1:
        st.info("No data available for year list progress.")
    else:
        default_yl = yl_years_avail[-2:] if len(yl_years_avail) >= 2 else yl_years_avail
        yl_years = st.multiselect(
            "Years to compare", yl_years_avail,
            default=default_yl, key="yl_years",
        )
        if not yl_years:
            st.info("Select at least one year.")
        else:
            yl_df = yl_df[yl_df["year"].isin(yl_years)].copy()
            yl_df["doy"] = yl_df["timestamp"].dt.dayofyear

            first_doy = yl_df.groupby(["year", "Com_Name"])["doy"].min().reset_index(name="First_DOY")

            cumul_rows = []
            for yr in sorted(yl_years):
                yr_data = first_doy[first_doy["year"] == yr].sort_values("First_DOY")
                for doy_val in range(1, 367):
                    count = (yr_data["First_DOY"] <= doy_val).sum()
                    cumul_rows.append({"Year": str(int(yr)), "Day_of_Year": doy_val, "Cumulative_Species": count})
            cumul_df = pd.DataFrame(cumul_rows)

            fig = px.line(
                cumul_df, x="Day_of_Year", y="Cumulative_Species",
                color="Year",
                title="Cumulative Species by Day of Year",
                labels={"Day_of_Year": "Day of year", "Cumulative_Species": "Cumulative species", "Year": "Year"},
                color_discrete_sequence=NATURE_PALETTE,
            )
            fig.update_traces(line=dict(width=2))
            st.plotly_chart(style_fig(fig), width="stretch")

    st.divider()

    # ── Detection Trends (Yearly / Monthly / Weekly) ──
    st.subheader("Detection Trends")

    trends_metric = st.radio(
        "Metric", ["Total Detections", "Unique Species"],
        horizontal=True, key="trends_metric",
    )
    _tm_species = trends_metric == "Unique Species"
    _tm_ylabel = "Species" if _tm_species else "Detections"

    def _trend_agg(df, groupby_cols):
        """Aggregate by size or nunique(Com_Name) based on metric toggle."""
        if _tm_species:
            return df.groupby(groupby_cols)["Com_Name"].nunique().reset_index(name="Count")
        return df.groupby(groupby_cols).size().reset_index(name="Count")

    def _trend_title(base):
        suffix = " \u00b7 Unique Species" if _tm_species else ""
        return f"{base}{suffix}"

    # Yearly
    yearly = filtered.dropna(subset=["timestamp"]).copy()
    yearly = _trend_agg(yearly, "year")
    yearly["year"] = yearly["year"].astype(int)
    fig = px.area(
        yearly, x="year", y="Count",
        title=_trend_title("Yearly Detection Trends"),
        labels={"year": "Year", "Count": _tm_ylabel},
    )
    fig.update_traces(
        line=dict(color=PRIMARY, width=2),
        fillcolor="rgba(61,107,68,0.14)",
        marker=dict(size=6, color=PRIMARY),
        mode="lines+markers",
    )
    fig.update_layout(xaxis=dict(dtick=1))
    st.plotly_chart(style_fig(fig), width="stretch")

    # Monthly & Weekly — with optional year comparison
    trends_df = filtered.dropna(subset=["timestamp"]).copy()
    trends_df["year"] = trends_df["timestamp"].dt.year.astype(int)
    trends_years_avail = sorted(trends_df["year"].dropna().unique())

    trends_cmp = st.checkbox("Compare years", value=False, key="trends_cmp_years")

    if trends_cmp and len(trends_years_avail) >= 2:
        default_trends_yrs = trends_years_avail[-2:] if len(trends_years_avail) >= 2 else trends_years_avail
        trends_years = st.multiselect(
            "Years to compare", trends_years_avail,
            default=default_trends_yrs, key="trends_years",
        )
        if not trends_years:
            st.info("Select at least one year.")
        else:
            t_df = trends_df[trends_df["year"].isin(trends_years)].copy()

            # Monthly by year
            monthly_yr = _trend_agg(t_df, ["year", "month"])
            monthly_yr["Year"] = monthly_yr["year"].astype(str)
            fig = px.line(
                monthly_yr, x="month", y="Count", color="Year",
                title=_trend_title("Monthly Detection Trends by Year"),
                labels={"month": "Month", "Count": _tm_ylabel, "Year": "Year"},
                color_discrete_sequence=NATURE_PALETTE,
                markers=True,
            )
            fig.update_traces(line=dict(width=2), marker=dict(size=5))
            fig.update_layout(xaxis=dict(
                dtick=1,
                tickmode="array",
                tickvals=list(MONTH_LABELS.keys()),
                ticktext=list(MONTH_LABELS.values()),
            ))
            st.plotly_chart(style_fig(fig), width="stretch")

            # Weekly by year
            weekly_yr = _trend_agg(t_df, ["year", "week"])
            weekly_yr["Year"] = weekly_yr["year"].astype(str)
            fig = px.line(
                weekly_yr, x="week", y="Count", color="Year",
                title=_trend_title("Weekly Detection Trends by Year"),
                labels={"week": "Week of year", "Count": _tm_ylabel, "Year": "Year"},
                color_discrete_sequence=NATURE_PALETTE,
                markers=True,
            )
            fig.update_traces(line=dict(width=2), marker=dict(size=4))
            st.plotly_chart(style_fig(fig), width="stretch")
    else:
        # Monthly — aggregated
        monthly = _trend_agg(filtered, "month")
        fig = px.area(
            monthly, x="month", y="Count",
            title=_trend_title("Monthly Detection Trends"),
            labels={"month": "Month", "Count": _tm_ylabel},
        )
        fig.update_traces(
            line=dict(color=TERTIARY, width=2),
            fillcolor="rgba(184,144,64,0.14)",
            marker=dict(size=6, color=TERTIARY),
            mode="lines+markers",
        )
        fig.update_layout(xaxis=dict(
            dtick=1,
            tickmode="array",
            tickvals=list(MONTH_LABELS.keys()),
            ticktext=list(MONTH_LABELS.values()),
        ))
        st.plotly_chart(style_fig(fig), width="stretch")

        # Weekly — aggregated
        weekly = _trend_agg(filtered, "week")
        fig = px.area(
            weekly, x="week", y="Count",
            title=_trend_title("Weekly Detection Trends"),
            labels={"week": "Week of year", "Count": _tm_ylabel},
        )
        fig.update_traces(
            line=dict(color=SECONDARY, width=2),
            fillcolor="rgba(74,112,144,0.14)",
            marker=dict(size=5, color=SECONDARY),
            mode="lines+markers",
        )
        st.plotly_chart(style_fig(fig), width="stretch")

# ── Community ──────────────────────────────────────────────────────
elif page == "Community":

    def tod_chart(data: pd.DataFrame, title: str, by_species: bool, by_status: bool):
        """Render one Activity by Hour chart."""
        if len(data) == 0:
            st.warning("No data for this selection.")
            return
        if by_species:
            top_sp = data["Com_Name"].value_counts().head(20).index.tolist()
            tod_df = data[data["Com_Name"].isin(top_sp)].copy()
            sp_hour = tod_df.groupby(["hour", "Com_Name"]).size().reset_index(name="Count")
            sp_color_map = {
                sp: NATURE_PALETTE[i % len(NATURE_PALETTE)]
                for i, sp in enumerate(top_sp)
            }
            fig = px.area(
                sp_hour, x="hour", y="Count",
                color="Com_Name",
                title=title,
                labels={"hour": "Hour of day", "Count": "Detections", "Com_Name": "Species"},
                category_orders={"Com_Name": top_sp},
                color_discrete_map=sp_color_map,
            )
            fig.update_layout(xaxis=dict(dtick=1))
            fig.update_traces(marker_line_width=0)
            hourly = data.groupby("hour").size().reset_index(name="Count")
            fig.add_scatter(
                x=hourly["hour"], y=hourly["Count"],
                mode="lines+markers",
                line=dict(color="#1a2416", width=2.5, dash="dot"),
                marker=dict(size=5, color="#1a2416"),
                name="Total", showlegend=True,
            )
        elif by_status:
            status_hour = (
                data.groupby(["hour", "UK_Status"])
                .size()
                .reset_index(name="Count")
            )
            cmap = status_color_map(status_hour["UK_Status"].unique())
            fig = px.line(
                status_hour,
                x="hour", y="Count",
                color="UK_Status",
                markers=True,
                title=title,
                labels={"hour": "Hour of day", "Count": "Detections", "UK_Status": "UK Status"},
                color_discrete_map=cmap,
            )
            fig.update_layout(xaxis=dict(dtick=1))
            fig.update_traces(line=dict(width=2), marker=dict(size=5))
        else:
            hourly = data.groupby("hour").size().reset_index(name="Count")
            fig = px.area(
                hourly, x="hour", y="Count",
                title=title,
                labels={"hour": "Hour of day", "Count": "Detections"},
            )
            fig.update_traces(
                line=dict(color=PRIMARY, width=2),
                fillcolor="rgba(61,107,68,0.14)",
                marker=dict(size=5, color=PRIMARY),
                mode="lines+markers",
            )
            fig.update_layout(xaxis=dict(dtick=1))
        st.plotly_chart(style_fig(fig), width="stretch")

    # Controls row
    def _tod_on_months():
        if st.session_state.get("tod_cmp_months"):
            st.session_state["tod_cmp_seasons"] = False

    def _tod_on_seasons():
        if st.session_state.get("tod_cmp_seasons"):
            st.session_state["tod_cmp_months"] = False

    def _tod_on_species():
        if st.session_state.get("tod_by_species"):
            st.session_state["tod_by_status"] = False

    def _tod_on_status():
        if st.session_state.get("tod_by_status"):
            st.session_state["tod_by_species"] = False

    tc1, tc2, tc3, tc4 = st.columns(4, gap="large")
    with tc1:
        show_by_species = st.checkbox("Show by species", value=False, key="tod_by_species", on_change=_tod_on_species)
    with tc2:
        show_by_status = st.checkbox("Show by status", value=False, key="tod_by_status", on_change=_tod_on_status)
    with tc3:
        tod_cmp_months = st.checkbox("Compare two months", value=False, key="tod_cmp_months", on_change=_tod_on_months)
    with tc4:
        tod_cmp_seasons = st.checkbox("Compare two seasons", value=False, key="tod_cmp_seasons", on_change=_tod_on_seasons)

    # Pre-filter base for compare overrides
    _tod_base = _filtered_pre_season_month.dropna(subset=["timestamp"]).copy()
    if exclude_review:
        _tod_base = _tod_base[_tod_base["UK_Status"] != "Review Recording"].copy()

    years_label  = "All years" if year_mode == "All years" else ", ".join(map(str, selected_years))
    season_label = "All seasons" if selected_season == "All" else selected_season
    month_label  = "All months" if month_mode == "All months" else chosen_month

    if tod_cmp_months:
        left, right = st.columns(2, gap="large")
        with left:
            tod_ma = st.selectbox("Month A", month_names_list, index=5, key="tod_month_a")
        with right:
            tod_mb = st.selectbox("Month B", month_names_list, index=6, key="tod_month_b")

        _cmp = _tod_base.copy()
        if selected_season != "All":
            _cmp = _cmp[_cmp["season"] == selected_season].copy()

        l, r = st.columns(2, gap="large")
        with l:
            tod_chart(_cmp[_cmp["month_num"] == month_num_by_name[tod_ma]],
                      f"{tod_ma} · {years_label} · {season_label}", show_by_species, show_by_status)
        with r:
            tod_chart(_cmp[_cmp["month_num"] == month_num_by_name[tod_mb]],
                      f"{tod_mb} · {years_label} · {season_label}", show_by_species, show_by_status)

    elif tod_cmp_seasons:
        seasons_opts = ["Spring", "Summer", "Autumn", "Winter"]
        left, right = st.columns(2, gap="large")
        with left:
            tod_sa = st.selectbox("Season A", seasons_opts, index=0, key="tod_season_a")
        with right:
            tod_sb = st.selectbox("Season B", seasons_opts, index=1, key="tod_season_b")

        _cmp = _tod_base.copy()
        if month_mode == "Choose month" and chosen_month:
            _cmp = _cmp[_cmp["month_num"] == month_num_by_name[chosen_month]].copy()

        l, r = st.columns(2, gap="large")
        with l:
            tod_chart(_cmp[_cmp["season"] == tod_sa],
                      f"{tod_sa} · {years_label} · {month_label}", show_by_species, show_by_status)
        with r:
            tod_chart(_cmp[_cmp["season"] == tod_sb],
                      f"{tod_sb} · {years_label} · {month_label}", show_by_species, show_by_status)

    else:
        tod_chart(filtered, f"Activity by Hour · {month_label} · {years_label} · {season_label}",
                  show_by_species, show_by_status)

    st.divider()

    # ── Heatmap ──
    st.subheader("Activity Heatmap")

    heatmap_data = (
        filtered.groupby(["month", "hour"])
        .size()
        .reset_index(name="Count")
    )
    # Pivot to a full 12×24 grid so every month gets its own row
    heatmap_pivot = heatmap_data.pivot(index="month", columns="hour", values="Count").fillna(0)
    heatmap_pivot = heatmap_pivot.reindex(index=range(1, 13), columns=range(24), fill_value=0)

    fig = px.imshow(
        heatmap_pivot.values,
        x=list(range(24)),
        y=list(MONTH_LABELS.values()),
        title="Activity Heatmap · Hour vs Month",
        color_continuous_scale=HEATMAP_SCALE,
        labels={"x": "Hour of day", "y": "Month", "color": "Detections"},
        aspect="auto",
    )
    fig.update_layout(
        xaxis=dict(dtick=1),
        yaxis=dict(dtick=1),
        coloraxis_colorbar=dict(
            title="Detections",
            tickfont=dict(size=11, color="#4a5c44"),
            title_font=dict(size=12, color="#4a5c44"),
            thickness=14,
        ),
    )
    st.plotly_chart(style_fig(fig), width="stretch")

    st.divider()

    st.subheader("Community Composition by Hour (Top 20 species, %)")

    comp_df = filtered.dropna(subset=["timestamp"]).copy()

    # Build colour map from the broadest dataset (pre-season/month filter)
    # so compare modes can look up any species that exists in the data.
    _comp_base_all = _filtered_pre_season_month.dropna(subset=["timestamp"]).copy()
    if exclude_review:
        _comp_base_all = _comp_base_all[_comp_base_all["UK_Status"] != "Review Recording"].copy()
    _species_ranked = _comp_base_all["Com_Name"].value_counts().index.tolist()
    species_color_map = {
        sp: NATURE_PALETTE[i % len(NATURE_PALETTE)]
        for i, sp in enumerate(_species_ranked)
    }

    seasons_list = ["Spring", "Summer", "Autumn", "Winter"]

    # Mutually exclusive checkboxes: toggling one clears the other via callbacks
    # that fire before the next render, avoiding Streamlit's "can't set widget
    # state after creation" error.
    def _on_months_change():
        if st.session_state.get("cc_cmp_months"):
            st.session_state["cc_cmp_seasons"] = False

    def _on_seasons_change():
        if st.session_state.get("cc_cmp_seasons"):
            st.session_state["cc_cmp_months"] = False

    c_cmp_m, c_cmp_s = st.columns(2, gap="large")
    with c_cmp_m:
        compare_months = st.checkbox(
            "Compare two months", value=False,
            key="cc_cmp_months", on_change=_on_months_change,
        )
    with c_cmp_s:
        compare_seasons = st.checkbox(
            "Compare two seasons", value=False,
            key="cc_cmp_seasons", on_change=_on_seasons_change,
        )

    def composition_plot(df_in: pd.DataFrame, title: str):
        if len(df_in) == 0:
            st.warning("No data for this selection.")
            return

        top_species = df_in["Com_Name"].value_counts().head(20).index.tolist()
        df_in = df_in[df_in["Com_Name"].isin(top_species)].copy()

        comp_hour = (
            df_in.groupby(["hour", "Com_Name"])
            .size()
            .reset_index(name="Count")
        )
        comp_hour["Percent"] = (
            comp_hour.groupby("hour")["Count"]
            .transform(lambda x: (x / x.sum()) * 100)
        )

        color_map = {
            sp: species_color_map.get(sp, NATURE_PALETTE[i % len(NATURE_PALETTE)])
            for i, sp in enumerate(top_species)
        }

        fig = px.bar(
            comp_hour,
            x="hour", y="Percent",
            color="Com_Name",
            title=title,
            labels={"hour": "Hour of day", "Percent": "% of detections", "Com_Name": "Species"},
            category_orders={"Com_Name": top_species},
            color_discrete_map=color_map,
        )
        fig.update_layout(barmode="stack", xaxis=dict(dtick=1))
        fig.update_traces(marker_line_width=0)
        st.plotly_chart(style_fig(fig), width="stretch")

    # Build label parts from global sidebar state
    years_label  = "All years" if year_mode == "All years" else ", ".join(map(str, selected_years))
    season_label = "All seasons" if selected_season == "All" else selected_season
    month_label  = "All months" if month_mode == "All months" else chosen_month

    # For compare modes, use pre-season/month filtered data so the compare
    # selectors override the global season/month filters for this page.
    _comp_base = _filtered_pre_season_month.dropna(subset=["timestamp"]).copy()
    if exclude_review:
        _comp_base = _comp_base[_comp_base["UK_Status"] != "Review Recording"].copy()

    if compare_months:
        left, right = st.columns(2, gap="large")
        with left:
            month_a = st.selectbox("Month A", month_names_list, index=5, key="cc_month_a")
        with right:
            month_b = st.selectbox("Month B", month_names_list, index=6, key="cc_month_b")

        # Apply global season filter but override month
        _cmp_df = _comp_base.copy()
        if selected_season != "All":
            _cmp_df = _cmp_df[_cmp_df["season"] == selected_season].copy()

        df_a = _cmp_df[_cmp_df["month_num"] == month_num_by_name[month_a]].copy()
        df_b = _cmp_df[_cmp_df["month_num"] == month_num_by_name[month_b]].copy()

        l, r = st.columns(2, gap="large")
        with l:
            composition_plot(df_a, f"{month_a} · {years_label} · {season_label}")
        with r:
            composition_plot(df_b, f"{month_b} · {years_label} · {season_label}")

    elif compare_seasons:
        left, right = st.columns(2, gap="large")
        with left:
            season_a = st.selectbox("Season A", seasons_list, index=0, key="cc_season_a")
        with right:
            season_b = st.selectbox("Season B", seasons_list, index=1, key="cc_season_b")

        # Apply global month filter but override season
        _cmp_df = _comp_base.copy()
        if month_mode == "Choose month" and chosen_month:
            _cmp_df = _cmp_df[_cmp_df["month_num"] == month_num_by_name[chosen_month]].copy()

        df_a = _cmp_df[_cmp_df["season"] == season_a].copy()
        df_b = _cmp_df[_cmp_df["season"] == season_b].copy()

        l, r = st.columns(2, gap="large")
        with l:
            composition_plot(df_a, f"{season_a} · {years_label} · {month_label}")
        with r:
            composition_plot(df_b, f"{season_b} · {years_label} · {month_label}")

    elif len(comp_df) == 0:
        st.info("No data available for the selected filters.")
    else:
        composition_plot(comp_df, f"{month_label} · {years_label} · {season_label}")

    st.divider()

    # ── Status / Diet Over Time ──
    st.subheader("Composition Over Time")

    comp_mode = st.radio("Breakdown by", ["UK Status", "Diet"], horizontal=True, key="comp_over_time")

    tmp = filtered.dropna(subset=["timestamp"]).copy()

    if comp_mode == "UK Status":
        comp_col = "UK_Status"
        comp_label = "UK Status"
        comp_month = (
            tmp.groupby(["month", comp_col])
            .size()
            .reset_index(name="Count")
        )
        comp_month["Percent"] = (
            comp_month.groupby("month")["Count"]
            .transform(lambda x: (x / x.sum()) * 100)
        )
        cmap = status_color_map(comp_month[comp_col].unique())
        fig = px.area(
            comp_month,
            x="month", y="Percent",
            color=comp_col,
            title="Monthly Status Composition",
            labels={"month": "Month", "Percent": "% of detections", comp_col: comp_label},
            color_discrete_map=cmap,
        )
    else:
        comp_col = "Diet"
        comp_label = "Diet"
        comp_month = (
            tmp.groupby(["month", comp_col])
            .size()
            .reset_index(name="Count")
        )
        comp_month["Percent"] = (
            comp_month.groupby("month")["Count"]
            .transform(lambda x: (x / x.sum()) * 100)
        )
        fig = px.area(
            comp_month,
            x="month", y="Percent",
            color=comp_col,
            title="Monthly Diet Composition",
            labels={"month": "Month", "Percent": "% of detections", comp_col: comp_label},
            color_discrete_map=DIET_COLORS,
        )

    fig.update_layout(xaxis=dict(
        dtick=1,
        tickmode="array",
        tickvals=list(MONTH_LABELS.keys()),
        ticktext=list(MONTH_LABELS.values()),
    ))
    st.plotly_chart(style_fig(fig), width="stretch")

    st.divider()

    # ── Species Co-occurrence ──
    st.subheader("Species Co-occurrence")

    co_topn = st.slider("Top N species", 5, 20, 15, key="co_topn")
    co_unit = st.radio("Co-occurrence unit", ["Day", "Hour"], horizontal=True, key="co_unit")

    co_df = filtered.dropna(subset=["timestamp"]).copy()

    if len(co_df) == 0:
        st.info("No data available for co-occurrence analysis.")
    else:
        top_co = co_df["Com_Name"].value_counts().head(co_topn).index.tolist()
        co_df = co_df[co_df["Com_Name"].isin(top_co)].copy()

        if co_unit == "Day":
            co_df["unit"] = co_df["timestamp"].dt.date.astype(str)
        else:
            co_df["unit"] = co_df["timestamp"].dt.strftime("%Y-%m-%d-%H")

        presence = co_df.groupby(["unit", "Com_Name"]).size().unstack(fill_value=0)
        presence = (presence > 0).astype(int)
        # Ensure all top species are columns
        for sp in top_co:
            if sp not in presence.columns:
                presence[sp] = 0
        presence = presence[top_co]

        dot = presence.T.values @ presence.values  # species x species
        counts = presence.sum(axis=0).values
        min_counts = np.minimum(counts[:, None], counts[None, :])
        min_counts[min_counts == 0] = 1  # avoid division by zero
        norm_co = dot / min_counts
        np.fill_diagonal(norm_co, 0)

        fig = px.imshow(
            norm_co,
            x=top_co, y=top_co,
            title=f"Species Co-occurrence (normalised, by {co_unit.lower()})",
            color_continuous_scale=HEATMAP_SCALE,
            labels={"color": "Co-occurrence"},
            aspect="auto",
        )
        fig.update_layout(
            coloraxis_colorbar=dict(
                title="Co-occurrence",
                tickfont=dict(size=11, color="#4a5c44"),
                title_font=dict(size=12, color="#4a5c44"),
                thickness=14,
            ),
        )
        st.plotly_chart(style_fig(fig), width="stretch")

    st.divider()

    # ── Diversity Indices ──
    st.subheader("Diversity Indices")

    div_df = filtered.dropna(subset=["timestamp"]).copy()
    div_df["year"] = div_df["timestamp"].dt.year.astype(int)
    div_years_avail = sorted(div_df["year"].dropna().unique())

    div_cmp = st.checkbox("Compare years", value=False, key="div_cmp_years")

    if len(div_df) == 0:
        st.info("No data available for diversity index computation.")
    elif div_cmp and len(div_years_avail) >= 2:
        default_div_yrs = div_years_avail[-2:] if len(div_years_avail) >= 2 else div_years_avail
        div_years = st.multiselect(
            "Years to compare", div_years_avail,
            default=default_div_yrs, key="div_years",
        )
        if not div_years:
            st.info("Select at least one year.")
        else:
            d_df = div_df[div_df["year"].isin(div_years)].copy()

            # Compute indices per year × month
            div_rows = []
            for yr in sorted(div_years):
                for m in range(1, 13):
                    p_df = d_df[(d_df["year"] == yr) & (d_df["month"] == m)]
                    counts = p_df["Com_Name"].value_counts().values
                    total = counts.sum()
                    richness = len(counts)
                    if total > 0 and richness > 0:
                        proportions = counts / total
                        shannon = -np.sum(proportions * np.log(proportions))
                        simpson = 1 - np.sum(proportions ** 2)
                    else:
                        shannon = 0.0
                        simpson = 0.0
                    div_rows.append({"Year": str(yr), "month": m, "Shannon_H": shannon,
                                     "Simpson_1D": simpson, "Unique_Species": richness})
            div_result = pd.DataFrame(div_rows)

            _month_tick = dict(dtick=1, tickmode="array",
                               tickvals=list(MONTH_LABELS.keys()),
                               ticktext=list(MONTH_LABELS.values()))

            fig_h = px.line(
                div_result, x="month", y="Shannon_H", color="Year",
                title="Shannon Diversity (H') by Year",
                labels={"month": "Month", "Shannon_H": "H'", "Year": "Year"},
                color_discrete_sequence=NATURE_PALETTE,
                markers=True,
            )
            fig_h.update_traces(line=dict(width=2), marker=dict(size=5))
            fig_h.update_layout(xaxis=_month_tick)
            st.plotly_chart(style_fig(fig_h), width="stretch")

            fig_s = px.line(
                div_result, x="month", y="Simpson_1D", color="Year",
                title="Simpson's Diversity (1-D) by Year",
                labels={"month": "Month", "Simpson_1D": "1-D", "Year": "Year"},
                color_discrete_sequence=NATURE_PALETTE,
                markers=True,
            )
            fig_s.update_traces(line=dict(width=2), marker=dict(size=5))
            fig_s.update_layout(xaxis=_month_tick)
            st.plotly_chart(style_fig(fig_s), width="stretch")

            fig_r = px.line(
                div_result, x="month", y="Unique_Species", color="Year",
                title="Unique Species per Month by Year",
                labels={"month": "Month", "Unique_Species": "Unique species", "Year": "Year"},
                color_discrete_sequence=NATURE_PALETTE,
                markers=True,
            )
            fig_r.update_traces(line=dict(width=2), marker=dict(size=5))
            fig_r.update_layout(xaxis=_month_tick)
            st.plotly_chart(style_fig(fig_r), width="stretch")
    else:
        div_res = st.radio("Time resolution", ["Month", "Week"], horizontal=True, key="div_res")

        if div_res == "Month":
            div_df["period"] = div_df["timestamp"].dt.to_period("M").astype(str)
        else:
            div_df["period"] = (
                div_df["timestamp"].dt.isocalendar().year.astype(str) + "-W"
                + div_df["timestamp"].dt.isocalendar().week.astype(str).str.zfill(2)
            )

        periods = sorted(div_df["period"].unique())
        div_rows = []
        for p in periods:
            p_df = div_df[div_df["period"] == p]
            counts = p_df["Com_Name"].value_counts().values
            total = counts.sum()
            richness = len(counts)
            if total > 0 and richness > 0:
                proportions = counts / total
                shannon = -np.sum(proportions * np.log(proportions))
                simpson = 1 - np.sum(proportions ** 2)
            else:
                shannon = 0.0
                simpson = 0.0
            div_rows.append({"Period": p, "Shannon_H": shannon, "Simpson_1D": simpson, "Unique_Species": richness})
        div_result = pd.DataFrame(div_rows)

        fig_h = px.line(
            div_result, x="Period", y="Shannon_H",
            title="Shannon Diversity (H')",
            labels={"Period": div_res, "Shannon_H": "H'"},
            markers=True,
        )
        fig_h.update_traces(line=dict(color=PRIMARY, width=2), marker=dict(size=5, color=PRIMARY))
        st.plotly_chart(style_fig(fig_h), width="stretch")

        fig_s = px.line(
            div_result, x="Period", y="Simpson_1D",
            title="Simpson's Diversity (1-D)",
            labels={"Period": div_res, "Simpson_1D": "1-D"},
            markers=True,
        )
        fig_s.update_traces(line=dict(color=SECONDARY, width=2), marker=dict(size=5, color=SECONDARY))
        st.plotly_chart(style_fig(fig_s), width="stretch")

        fig_r = px.line(
            div_result, x="Period", y="Unique_Species",
            title="Unique Species per Month",
            labels={"Period": div_res, "Unique_Species": "Unique species"},
            markers=True,
        )
        fig_r.update_traces(line=dict(color=TERTIARY, width=2), marker=dict(size=5, color=TERTIARY))
        st.plotly_chart(style_fig(fig_r), width="stretch")

# ── NMDS ──────────────────────────────────────────────────────────────────
elif page == "NMDS":
    st.subheader("NMDS Ordination")

    nmds_c1, nmds_c2, nmds_c3 = st.columns(3)
    with nmds_c1:
        nmds_matrix = st.selectbox(
            "Feature matrix",
            ["Species × Peak Activity Time", "Species × Month", "Species × Season"],
            key="nmds_matrix",
        )
    with nmds_c2:
        nmds_colour = st.selectbox(
            "Colour by",
            ["Diet", "UK Status", "Peak Activity Time", "Peak Season"],
            key="nmds_colour",
        )
    with nmds_c3:
        nmds_min_det = st.slider(
            "Minimum detections per species", 1, 100, 5, key="nmds_min_det",
        )

    # Filter to species with enough detections
    nmds_det_counts = filtered["Com_Name"].value_counts()
    nmds_valid_species = nmds_det_counts[nmds_det_counts >= nmds_min_det].index.tolist()
    nmds_df = filtered[filtered["Com_Name"].isin(nmds_valid_species)].copy()

    if len(nmds_valid_species) < 5:
        st.warning(
            f"Only {len(nmds_valid_species)} species meet the minimum detection threshold. "
            "At least 5 are needed for NMDS. Try lowering the threshold or broadening filters."
        )
    else:
        nmds_ts = nmds_df.dropna(subset=["timestamp"]).copy()

        # Build pivot table based on chosen matrix
        _season_map = {1: "Winter", 2: "Winter", 3: "Spring", 4: "Spring", 5: "Spring",
                       6: "Summer", 7: "Summer", 8: "Summer", 9: "Autumn", 10: "Autumn",
                       11: "Autumn", 12: "Winter"}
        if nmds_matrix == "Species × Peak Activity Time":
            nmds_ts["_unit"] = nmds_ts["hour"].apply(assign_time_bucket)
            all_cols = list(TIME_BUCKET_COLORS.keys())
        elif nmds_matrix == "Species × Month":
            nmds_ts["_unit"] = nmds_ts["month"].map(MONTH_LABELS)
            all_cols = list(MONTH_LABELS.values())
        else:  # Species × Season
            nmds_ts["_unit"] = nmds_ts["month"].map(_season_map)
            all_cols = list(SEASON_COLORS.keys())

        nmds_pivot = nmds_ts.pivot_table(
            index="Com_Name", columns="_unit", values="timestamp",
            aggfunc="count", fill_value=0,
        )
        # Ensure all columns present
        for c in all_cols:
            if c not in nmds_pivot.columns:
                nmds_pivot[c] = 0
        nmds_pivot = nmds_pivot[all_cols]

        # Normalise rows to proportions
        row_sums = nmds_pivot.sum(axis=1).replace(0, 1)
        nmds_norm = nmds_pivot.div(row_sums, axis=0)

        species_list = nmds_norm.index.tolist()
        coords, stress = compute_nmds(nmds_norm.values, tuple(species_list))

        # Build result DataFrame with metadata
        nmds_result = pd.DataFrame({
            "Species": species_list,
            "NMDS1": coords[:, 0],
            "NMDS2": coords[:, 1],
        })

        # Add metadata per species
        sp_meta = nmds_df.groupby("Com_Name").agg(
            Diet=("Diet", lambda x: x.mode().iloc[0] if len(x.mode()) else "Unclassified"),
            UK_Status=("UK_Status", lambda x: x.mode().iloc[0] if len(x.mode()) else "Unknown"),
            Detections=("Com_Name", "count"),
        ).reset_index().rename(columns={"Com_Name": "Species"})

        # Dominant time bucket
        nmds_ts["_tb"] = nmds_ts["hour"].apply(assign_time_bucket)
        tb_counts = nmds_ts.groupby(["Com_Name", "_tb"]).size().reset_index(name="n")
        dom_tb = tb_counts.loc[tb_counts.groupby("Com_Name")["n"].idxmax()][["Com_Name", "_tb"]]
        dom_tb.columns = ["Species", "Dominant_Time_Bucket"]

        # Peak season
        nmds_ts["_season"] = nmds_ts["month"].map(_season_map)
        season_counts = nmds_ts.groupby(["Com_Name", "_season"]).size().reset_index(name="n")
        peak_season = season_counts.loc[season_counts.groupby("Com_Name")["n"].idxmax()][["Com_Name", "_season"]]
        peak_season.columns = ["Species", "Peak_Season"]

        nmds_result = nmds_result.merge(sp_meta, on="Species", how="left")
        nmds_result = nmds_result.merge(dom_tb, on="Species", how="left")
        nmds_result = nmds_result.merge(peak_season, on="Species", how="left")

        # Select colour column and colour map
        if nmds_colour == "Diet":
            color_col = "Diet"
            color_map = DIET_COLORS
        elif nmds_colour == "UK Status":
            color_col = "UK_Status"
            color_map = STATUS_COLORS
        elif nmds_colour == "Peak Activity Time":
            color_col = "Dominant_Time_Bucket"
            color_map = TIME_BUCKET_COLORS
        else:  # Peak Season
            color_col = "Peak_Season"
            color_map = SEASON_COLORS

        fig_nmds = px.scatter(
            nmds_result,
            x="NMDS1", y="NMDS2",
            color=color_col,
            color_discrete_map=color_map,
            hover_name="Species",
            hover_data={
                "Diet": True,
                "UK_Status": True,
                "Dominant_Time_Bucket": True,
                "Peak_Season": True,
                "Detections": True,
                "NMDS1": ":.3f",
                "NMDS2": ":.3f",
            },
            title="NMDS — Species Similarity Ordination",
        )
        fig_nmds.update_traces(marker=dict(size=10, line=dict(width=1, color="rgba(26,36,22,0.3)")))

        # Draw convex hulls grouped by the dominant feature-matrix category
        # (i.e. which column of the matrix each species peaks in)
        from scipy.spatial import ConvexHull

        # Determine each species' dominant matrix column
        _dominant = nmds_pivot.idxmax(axis=1)  # Series: Com_Name -> column label
        _dominant.name = "_dominant_matrix_cat"
        nmds_result = nmds_result.merge(
            _dominant.reset_index().rename(columns={"Com_Name": "Species"}),
            on="Species", how="left",
        )

        # Pick hull colours and legend title based on the matrix type
        if nmds_matrix == "Species × Peak Activity Time":
            _hull_colors = TIME_BUCKET_COLORS
            _hull_title = "Peak activity time"
        elif nmds_matrix == "Species × Season":
            _hull_colors = SEASON_COLORS
            _hull_title = "Peak season"
        elif nmds_matrix == "Species × Month":
            # Generate month colours from the heatmap scale
            _month_greens = [
                "#c8dfa0", "#aed48a", "#94c974", "#7aaa6a", "#6b9e5e",
                "#5c8c5c", "#4d7a4d", "#3e683e", "#2d5233", "#3e683e",
                "#5c8c5c", "#94c974",
            ]
            _hull_colors = {m: _month_greens[i] for i, m in enumerate(MONTH_LABELS.values())}
            _hull_title = "Peak month"

        _hull_group_title = _hull_title
        for group_name, grp in nmds_result.groupby("_dominant_matrix_cat"):
            if len(grp) < 3:
                continue
            pts = grp[["NMDS1", "NMDS2"]].values
            try:
                hull = ConvexHull(pts)
            except Exception:
                continue
            hull_idx = list(hull.vertices) + [hull.vertices[0]]
            base_color = _hull_colors.get(group_name, "#8c9c8c")
            fig_nmds.add_trace(go.Scatter(
                x=pts[hull_idx, 0], y=pts[hull_idx, 1],
                mode="lines",
                fill="toself",
                fillcolor=f"rgba({_hex_to_rgb(base_color)}, 0.10)",
                line=dict(color=base_color, width=1.5, dash="dot"),
                name=f"{group_name}",
                legendgroup=f"hull_{group_name}",
                legendgrouptitle_text=_hull_group_title,
                showlegend=True,
                hoverinfo="skip",
            ))
            _hull_group_title = None

        st.plotly_chart(style_fig(fig_nmds), width="stretch")

        # Metrics row
        mc1, mc2, mc3 = st.columns(3)
        mc1.metric("Stress", f"{stress:.4f}")
        mc2.metric("Species", len(species_list))
        if stress < 0.05:
            quality = "Excellent"
        elif stress < 0.1:
            quality = "Good"
        elif stress < 0.2:
            quality = "Fair"
        else:
            quality = "Poor"
        mc3.metric("Stress Quality", quality)
        st.caption(
            "Stress measures how well the 2D layout preserves the original dissimilarities. "
            "Excellent < 0.05, Good < 0.1, Fair < 0.2, Poor ≥ 0.2."
        )

# ── Dawn Chorus Overview ──────────────────────────────────────────────────
elif page == "Dawn Chorus Overview":

    # ── Dawn Chorus Tracker ──
    st.subheader("Dawn Chorus Tracker")
    dc_c1, dc_c2 = st.columns(2, gap="large")
    with dc_c1:
        dc_topn = st.slider("Top N species", 5, 20, 12, key="dc_topn")
    with dc_c2:
        dc_time_mode = st.radio("Time format", ["Local (GMT/BST)", "UTC"], horizontal=True, key="dc_time_mode")

    dc_df = filtered.dropna(subset=["timestamp"]).copy()
    dc_df = dc_df[(dc_df["hour"] >= 3) & (dc_df["hour"] <= 10)]

    if len(dc_df) == 0:
        st.info("No detections in the dawn window (03:00-10:00) for current filters.")
    else:
        top_dawn = dc_df["Com_Name"].value_counts().head(dc_topn).index.tolist()
        dc_df = dc_df[dc_df["Com_Name"].isin(top_dawn)].copy()
        dc_df["date"] = dc_df["timestamp"].dt.date

        # Open-Meteo sunrise is in GMT (smooth). Detections are in local time.
        # In UTC mode: convert detections local→UTC so both align with sunrise.
        # In local mode: use raw detection hours; shift sunrise by +1 during BST.
        use_utc = dc_time_mode == "UTC"
        if use_utc:
            dc_df["decimal_hour"] = to_utc_hour(dc_df["timestamp"])
            hour_label = "Hour (UTC)"
        else:
            dc_df["decimal_hour"] = dc_df["timestamp"].dt.hour + dc_df["timestamp"].dt.minute / 60.0
            hour_label = "Hour (local)"

        earliest = (
            dc_df.groupby(["date", "Com_Name"])["decimal_hour"]
            .min()
            .reset_index(name="Earliest_Hour")
        )

        color_map = {
            sp: NATURE_PALETTE[i % len(NATURE_PALETTE)]
            for i, sp in enumerate(top_dawn)
        }
        fig = px.scatter(
            earliest, x="date", y="Earliest_Hour",
            color="Com_Name",
            title="Earliest Detection by Day (Dawn Window)",
            labels={"date": "Date", "Earliest_Hour": hour_label, "Com_Name": "Species"},
            color_discrete_map=color_map,
        )
        fig.update_traces(marker=dict(size=5, opacity=0.7))

        # Always show sunrise
        w_lat = float(dc_df["Lat"].mode().iloc[0])
        w_lon = float(dc_df["Lon"].mode().iloc[0])
        w_start = dc_df["timestamp"].min().strftime("%Y-%m-%d")
        w_end = dc_df["timestamp"].max().strftime("%Y-%m-%d")
        _, sunrise_daily = fetch_weather(w_lat, w_lon, w_start, w_end)
        if sunrise_daily is not None and "sunrise" in sunrise_daily.columns:
            sunrise_daily = sunrise_daily.copy()
            if use_utc:
                # Sunrise from Open-Meteo is already GMT/UTC
                sunrise_daily["sunrise_hour"] = (
                    sunrise_daily["sunrise"].dt.hour + sunrise_daily["sunrise"].dt.minute / 60.0
                )
            else:
                # Convert GMT sunrise to local time (add 1 hour during BST)
                sunrise_daily["sunrise_hour"] = sunrise_daily["sunrise"].apply(
                    lambda t: t.replace(tzinfo=_TZ_UTC).astimezone(_TZ_LONDON)
                ).dt.hour + sunrise_daily["sunrise"].apply(
                    lambda t: t.replace(tzinfo=_TZ_UTC).astimezone(_TZ_LONDON)
                ).dt.minute / 60.0
            fig.add_scatter(
                x=sunrise_daily["date"], y=sunrise_daily["sunrise_hour"],
                mode="lines", line=dict(color="#c47a5a", width=2.5, dash="dash"),
                name="Sunrise", showlegend=True,
            )

        st.plotly_chart(style_fig(fig), width="stretch")

    st.divider()

    # ── First Detection vs Sunrise ──
    st.subheader("First Detection vs Sunrise")

    fds_df = filtered.dropna(subset=["timestamp"]).copy()
    fds_df = fds_df[(fds_df["hour"] >= 3) & (fds_df["hour"] <= 10)].copy()

    if len(fds_df) == 0:
        st.info("No dawn detections (03:00-10:00) in the current filters.")
    else:
        fds_df["date"] = fds_df["timestamp"].dt.date

        # Reuse the same time mode from the dawn chorus tracker
        fds_utc = dc_time_mode == "UTC"

        # Earliest detection per day
        fds_earliest = (
            fds_df.groupby("date")["timestamp"]
            .min()
            .reset_index(name="earliest_detection")
        )
        if fds_utc:
            fds_earliest["earliest_hour"] = to_utc_hour(fds_earliest["earliest_detection"])
        else:
            fds_earliest["earliest_hour"] = (
                fds_earliest["earliest_detection"].dt.hour
                + fds_earliest["earliest_detection"].dt.minute / 60.0
            )

        # Fetch weather for sunrise data
        fds_lat = float(fds_df["Lat"].mode().iloc[0])
        fds_lon = float(fds_df["Lon"].mode().iloc[0])
        fds_start = fds_df["timestamp"].min().strftime("%Y-%m-%d")
        fds_end = fds_df["timestamp"].max().strftime("%Y-%m-%d")
        fds_weather_hourly, fds_weather_daily = fetch_weather(fds_lat, fds_lon, fds_start, fds_end)

        if fds_weather_daily is not None and "sunrise" in fds_weather_daily.columns:
            sunrise_df = fds_weather_daily[["date", "sunrise"]].copy()
            if fds_utc:
                sunrise_df["sunrise_hour"] = (
                    sunrise_df["sunrise"].dt.hour + sunrise_df["sunrise"].dt.minute / 60.0
                )
            else:
                _sr_local = sunrise_df["sunrise"].apply(
                    lambda t: t.replace(tzinfo=_TZ_UTC).astimezone(_TZ_LONDON)
                )
                sunrise_df["sunrise_hour"] = _sr_local.dt.hour + _sr_local.dt.minute / 60.0

            # Temperature at sunrise hour
            sunrise_temps = []
            for _, row in sunrise_df.iterrows():
                sr_row = fds_weather_daily[fds_weather_daily["date"] == row["date"]]
                if len(sr_row):
                    sr_hour = int(sr_row.iloc[0]["sunrise"].hour)
                else:
                    sr_hour = 6
                match = fds_weather_hourly[
                    (fds_weather_hourly["date"] == row["date"]) & (fds_weather_hourly["hour"] == sr_hour)
                ] if fds_weather_hourly is not None else pd.DataFrame()
                temp = match["temperature"].iloc[0] if len(match) > 0 else None
                sunrise_temps.append({"date": row["date"], "sunrise_hour": row["sunrise_hour"],
                                      "sunrise_temp": temp})
            sunrise_temp_df = pd.DataFrame(sunrise_temps)

            fds_merged = fds_earliest.merge(sunrise_temp_df, on="date", how="inner")
            fds_merged = fds_merged.dropna(subset=["sunrise_temp"])

            hour_suffix = "UTC" if fds_utc else "local"

            if len(fds_merged) == 0:
                st.info("No matching weather data for dawn detections.")
            else:
                d_l2, d_r2 = st.columns(2, gap="large")
                with d_l2:
                    fig = px.scatter(
                        fds_merged, x="sunrise_temp", y="earliest_hour",
                        title="First Detection vs Sunrise Temperature",
                        labels={"sunrise_temp": "Temperature at sunrise (°C)",
                                "earliest_hour": f"Earliest detection ({hour_suffix})"},
                        hover_data={"date": True},
                    )
                    fig.update_traces(marker=dict(size=8, color=PRIMARY, opacity=0.7))
                    if len(fds_merged) > 2:
                        z = np.polyfit(fds_merged["sunrise_temp"], fds_merged["earliest_hour"], 1)
                        x_range = np.linspace(fds_merged["sunrise_temp"].min(),
                                              fds_merged["sunrise_temp"].max(), 50)
                        fig.add_scatter(x=x_range, y=np.polyval(z, x_range),
                                        mode="lines", line=dict(color=TERTIARY, width=2, dash="dash"),
                                        name="Trend", showlegend=True)
                    st.plotly_chart(style_fig(fig), width="stretch")

                with d_r2:
                    fig = px.scatter(
                        fds_merged, x="sunrise_hour", y="earliest_hour",
                        title="First Detection vs Sunrise Time",
                        labels={"sunrise_hour": f"Sunrise ({hour_suffix})",
                                "earliest_hour": f"Earliest detection ({hour_suffix})"},
                        hover_data={"date": True},
                    )
                    fig.update_traces(marker=dict(size=8, color=SECONDARY, opacity=0.7))
                    # Add y=x reference line
                    xy_range = [min(fds_merged["sunrise_hour"].min(), fds_merged["earliest_hour"].min()),
                                max(fds_merged["sunrise_hour"].max(), fds_merged["earliest_hour"].max())]
                    fig.add_scatter(x=xy_range, y=xy_range,
                                    mode="lines", line=dict(color="#1a2416", width=1, dash="dot"),
                                    name="Sunrise = Detection", showlegend=True)
                    st.plotly_chart(style_fig(fig), width="stretch")
        else:
            st.info("Could not fetch sunrise data.")

# ── Weather & Activity ──────────────────────────────────────────────────────
elif page == "Weather & Activity":

    w_df = filtered.dropna(subset=["timestamp"]).copy()

    if len(w_df) == 0:
        st.info("No detection data available for weather analysis.")
    else:
        # Get location and date range from the data
        w_lat = float(w_df["Lat"].mode().iloc[0])
        w_lon = float(w_df["Lon"].mode().iloc[0])
        w_start = w_df["timestamp"].min().strftime("%Y-%m-%d")
        w_end = w_df["timestamp"].max().strftime("%Y-%m-%d")

        weather_hourly, weather_daily = fetch_weather(w_lat, w_lon, w_start, w_end)

        if weather_hourly is None or weather_daily is None:
            st.error("Could not fetch weather data from Open-Meteo.")
        else:
            # ── Prepare merged datasets ──
            w_df["date"] = w_df["timestamp"].dt.date
            w_df["hour"] = w_df["timestamp"].dt.hour

            # Daily detection counts
            daily_det = w_df.groupby("date").agg(
                det_count=("Com_Name", "size"),
                species_count=("Com_Name", "nunique"),
            ).reset_index()
            daily_merged = daily_det.merge(weather_daily, on="date", how="inner")

            # Hourly detection counts
            hourly_det = w_df.groupby(["date", "hour"]).size().reset_index(name="det_count")
            hourly_merged = hourly_det.merge(weather_hourly, on=["date", "hour"], how="inner")

            # ── 1. Detections vs Temperature ──
            st.subheader("Detections vs Temperature")

            fig = px.scatter(
                daily_merged, x="temp_max", y="det_count",
                color="precip_sum",
                color_continuous_scale=[[0, "#f5f3ee"], [0.5, "#6a90b0"], [1, "#4a5c70"]],
                title="Daily Detections vs Max Temperature",
                labels={"temp_max": "Max temperature (°C)", "det_count": "Detections",
                        "precip_sum": "Rainfall (mm)"},
                hover_data={"date": True},
            )
            fig.update_traces(marker=dict(size=8, opacity=0.7, line=dict(width=0.5, color="#1a2416")))
            # Add trendline manually
            if len(daily_merged) > 2:
                z = np.polyfit(daily_merged["temp_max"].dropna(), daily_merged.loc[daily_merged["temp_max"].notna(), "det_count"], 1)
                x_range = np.linspace(daily_merged["temp_max"].min(), daily_merged["temp_max"].max(), 50)
                fig.add_scatter(x=x_range, y=np.polyval(z, x_range),
                                mode="lines", line=dict(color=PRIMARY, width=2, dash="dash"),
                                name="Trend", showlegend=True)
            st.plotly_chart(style_fig(fig), width="stretch")

            st.divider()

            # ── 2. Rainy vs Dry Days ──
            st.subheader("Rainy vs Dry Days")

            # Merge hourly weather with hourly detections for activity profile
            rain_thresh = st.slider("Rain threshold (mm/day)", 0.0, 10.0, 1.0, key="w_rain_thresh")

            rain_days = set(weather_daily[weather_daily["precip_sum"] >= rain_thresh]["date"].tolist())
            w_df["day_type"] = w_df["date"].apply(lambda d: "Rainy" if d in rain_days else "Dry")

            rain_profile = w_df.groupby(["hour", "day_type"]).size().reset_index(name="Count")
            # Normalise by number of days of each type
            n_rain = max(len(rain_days & set(w_df["date"].unique())), 1)
            n_dry = max(len(set(w_df["date"].unique()) - rain_days), 1)
            rain_profile["Avg_Detections"] = rain_profile.apply(
                lambda r: r["Count"] / n_rain if r["day_type"] == "Rainy" else r["Count"] / n_dry,
                axis=1,
            )

            fig = px.line(
                rain_profile, x="hour", y="Avg_Detections", color="day_type",
                title=f"Average Hourly Activity: Rainy vs Dry Days (threshold: {rain_thresh}mm)",
                labels={"hour": "Hour of day", "Avg_Detections": "Avg detections per day",
                        "day_type": "Day type"},
                color_discrete_map={"Rainy": SECONDARY, "Dry": TERTIARY},
                markers=True,
            )
            fig.update_traces(line=dict(width=2), marker=dict(size=5))
            fig.update_layout(xaxis=dict(dtick=1))

            # Add KPI cards
            r_k1, r_k2, r_k3 = st.columns(3)
            r_k1.metric("Rainy days", f"{n_rain}")
            r_k2.metric("Dry days", f"{n_dry}")
            avg_rain_det = daily_merged[daily_merged["date"].isin(rain_days)]["det_count"].mean()
            avg_dry_det = daily_merged[~daily_merged["date"].isin(rain_days)]["det_count"].mean()
            r_k3.metric("Avg detections",
                        f"Rainy: {avg_rain_det:.0f}" if pd.notna(avg_rain_det) else "—",
                        f"Dry: {avg_dry_det:.0f}" if pd.notna(avg_dry_det) else None)

            st.plotly_chart(style_fig(fig), width="stretch")

            st.divider()

            # ── 3. Wind Speed Impact ──
            st.subheader("Wind Speed Impact")

            wind_bins = [0, 10, 20, 30, 100]
            wind_labels = ["Calm (0-10)", "Light (10-20)", "Moderate (20-30)", "Strong (30+)"]
            daily_merged["wind_bracket"] = pd.cut(
                daily_merged["wind_max"], bins=wind_bins, labels=wind_labels, right=False,
            )

            wind_agg = daily_merged.groupby("wind_bracket", observed=True).agg(
                avg_det=("det_count", "mean"),
                avg_species=("species_count", "mean"),
                day_count=("date", "count"),
            ).reset_index()

            w_l, w_r = st.columns(2, gap="large")
            with w_l:
                fig = px.bar(
                    wind_agg, x="wind_bracket", y="avg_det",
                    title="Avg Daily Detections by Wind Speed",
                    labels={"wind_bracket": "Wind speed (km/h)", "avg_det": "Avg detections"},
                    text="day_count",
                )
                fig.update_traces(marker_color=PRIMARY, marker_line_width=0,
                                  texttemplate="%{text} days", textposition="outside")
                st.plotly_chart(style_fig(fig), width="stretch")

            with w_r:
                fig = px.bar(
                    wind_agg, x="wind_bracket", y="avg_species",
                    title="Avg Species Richness by Wind Speed",
                    labels={"wind_bracket": "Wind speed (km/h)", "avg_species": "Avg unique species"},
                    text="day_count",
                )
                fig.update_traces(marker_color=SECONDARY, marker_line_width=0,
                                  texttemplate="%{text} days", textposition="outside")
                st.plotly_chart(style_fig(fig), width="stretch")

            st.divider()

            # ── 4. Weather Overlay on Monthly Trends ──
            st.subheader("Monthly Trends with Weather")

            daily_merged["month"] = pd.to_datetime(daily_merged["date"]).dt.month
            monthly_weather = daily_merged.groupby("month").agg(
                total_det=("det_count", "sum"),
                avg_temp=("temp_max", "mean"),
                total_rain=("precip_sum", "sum"),
            ).reset_index()

            fig = make_subplots(specs=[[{"secondary_y": True}]])
            fig.add_trace(
                go.Bar(x=monthly_weather["month"], y=monthly_weather["total_det"],
                       name="Detections", marker_color=PRIMARY, opacity=0.7),
                secondary_y=False,
            )
            fig.add_trace(
                go.Scatter(x=monthly_weather["month"], y=monthly_weather["avg_temp"],
                           name="Avg max temp (°C)", mode="lines+markers",
                           line=dict(color="#c47a5a", width=2.5),
                           marker=dict(size=7, color="#c47a5a")),
                secondary_y=True,
            )
            fig.update_layout(
                title="Monthly Detections & Temperature",
                xaxis=dict(dtick=1, tickmode="array",
                           tickvals=list(MONTH_LABELS.keys()),
                           ticktext=list(MONTH_LABELS.values())),
                legend=dict(x=0.01, y=0.99),
            )
            fig.update_yaxes(title_text="Detections", secondary_y=False)
            fig.update_yaxes(title_text="Temperature (°C)", secondary_y=True)
            st.plotly_chart(style_fig(fig), width="stretch")

            # Rainfall overlay
            fig2 = make_subplots(specs=[[{"secondary_y": True}]])
            fig2.add_trace(
                go.Bar(x=monthly_weather["month"], y=monthly_weather["total_det"],
                       name="Detections", marker_color=PRIMARY, opacity=0.7,
                       offsetgroup=0),
                secondary_y=False,
            )
            fig2.add_trace(
                go.Bar(x=monthly_weather["month"], y=monthly_weather["total_rain"],
                       name="Total rainfall (mm)", marker_color=SECONDARY, opacity=0.7,
                       offsetgroup=1),
                secondary_y=True,
            )
            fig2.update_layout(
                title="Monthly Detections & Rainfall",
                xaxis=dict(dtick=1, tickmode="array",
                           tickvals=list(MONTH_LABELS.keys()),
                           ticktext=list(MONTH_LABELS.values())),
                barmode="group",
                legend=dict(x=0.01, y=0.99),
            )
            fig2.update_yaxes(title_text="Detections", secondary_y=False)
            fig2.update_yaxes(title_text="Rainfall (mm)", secondary_y=True)
            st.plotly_chart(style_fig(fig2), width="stretch")

            st.divider()

            # ── 5. Species Diversity vs Conditions ──
            st.subheader("Species Diversity vs Conditions")

            d_l, d_r = st.columns(2, gap="large")
            with d_l:
                fig = px.scatter(
                    daily_merged, x="temp_max", y="species_count",
                    title="Unique Species vs Temperature",
                    labels={"temp_max": "Max temperature (°C)", "species_count": "Unique species"},
                    color="precip_sum",
                    color_continuous_scale=[[0, "#f5f3ee"], [0.5, "#6a90b0"], [1, "#4a5c70"]],
                )
                fig.update_traces(marker=dict(size=8, opacity=0.7))
                if len(daily_merged) > 2:
                    mask = daily_merged["temp_max"].notna()
                    z = np.polyfit(daily_merged.loc[mask, "temp_max"], daily_merged.loc[mask, "species_count"], 1)
                    x_range = np.linspace(daily_merged["temp_max"].min(), daily_merged["temp_max"].max(), 50)
                    fig.add_scatter(x=x_range, y=np.polyval(z, x_range),
                                    mode="lines", line=dict(color=PRIMARY, width=2, dash="dash"),
                                    name="Trend", showlegend=True)
                st.plotly_chart(style_fig(fig), width="stretch")

            with d_r:
                fig = px.scatter(
                    daily_merged, x="wind_max", y="species_count",
                    title="Unique Species vs Wind Speed",
                    labels={"wind_max": "Max wind speed (km/h)", "species_count": "Unique species"},
                    color="precip_sum",
                    color_continuous_scale=[[0, "#f5f3ee"], [0.5, "#6a90b0"], [1, "#4a5c70"]],
                )
                fig.update_traces(marker=dict(size=8, opacity=0.7))
                if len(daily_merged) > 2:
                    mask = daily_merged["wind_max"].notna()
                    z = np.polyfit(daily_merged.loc[mask, "wind_max"], daily_merged.loc[mask, "species_count"], 1)
                    x_range = np.linspace(daily_merged["wind_max"].min(), daily_merged["wind_max"].max(), 50)
                    fig.add_scatter(x=x_range, y=np.polyval(z, x_range),
                                    mode="lines", line=dict(color=PRIMARY, width=2, dash="dash"),
                                    name="Trend", showlegend=True)
                st.plotly_chart(style_fig(fig), width="stretch")


# ── Data Quality ─────────────────────────────────────────────────
elif page == "Data Quality":

    # ── Confidence Distribution ──
    st.subheader("Confidence Distribution")

    cd_topn = st.slider("Top N species", 5, 30, 20, key="cd_topn")
    cd_box = st.checkbox("Overlay box plot", value=True, key="cd_box")

    cd_df = filtered.dropna(subset=["Confidence"]).copy()

    if len(cd_df) == 0:
        st.info("No confidence data available.")
    else:
        # Sort species by median confidence
        medians = cd_df.groupby("Com_Name")["Confidence"].median().sort_values()
        top_cd = medians.tail(cd_topn).index.tolist()
        cd_df = cd_df[cd_df["Com_Name"].isin(top_cd)].copy()
        # Reorder by median
        species_order = medians.loc[medians.index.isin(top_cd)].index.tolist()

        color_map = {
            sp: NATURE_PALETTE[i % len(NATURE_PALETTE)]
            for i, sp in enumerate(species_order)
        }

        fig = px.violin(
            cd_df, x="Confidence", y="Com_Name",
            orientation="h",
            title="Confidence Distribution by Species",
            labels={"Confidence": "Confidence", "Com_Name": "Species"},
            color="Com_Name",
            color_discrete_map=color_map,
            category_orders={"Com_Name": species_order},
        )
        if cd_box:
            fig.update_traces(box_visible=True)
        fig.update_layout(showlegend=False, height=max(400, len(species_order) * 28))
        st.plotly_chart(style_fig(fig), width="stretch")

    st.divider()

    # ── False Positive Candidates ──
    st.subheader("False Positive Candidates")

    fp_thresh = st.slider("Confidence threshold", 0.0, 1.0, 0.7, key="fp_thresh")
    all_statuses = sorted(filtered["UK_Status"].dropna().unique())
    fp_default = [s for s in ["Rare vagrant", "Scarce visitor"] if s in all_statuses]
    fp_statuses = st.multiselect(
        "UK statuses to flag", all_statuses,
        default=fp_default, key="fp_statuses",
    )

    fp_df = filtered[filtered["Confidence"] <= fp_thresh].copy()
    if fp_statuses:
        fp_df = fp_df[fp_df["UK_Status"].isin(fp_statuses)].copy()

    if len(fp_df) == 0:
        st.info("No false positive candidates for current filters.")
    else:
        fp_left, fp_right = st.columns(2, gap="large")

        with fp_left:
            cmap = status_color_map(fp_df["UK_Status"].unique())
            fig = px.scatter(
                fp_df, x="Confidence", y="Com_Name",
                color="UK_Status",
                title="Low-Confidence Detections",
                labels={"Confidence": "Confidence", "Com_Name": "Species", "UK_Status": "UK Status"},
                color_discrete_map=cmap,
            )
            fig.update_traces(marker=dict(size=6, opacity=0.7))
            st.plotly_chart(style_fig(fig), width="stretch")

        with fp_right:
            summary = (
                fp_df.groupby(["Com_Name", "UK_Status"])
                .agg(Count=("Confidence", "size"), Avg_Confidence=("Confidence", "mean"))
                .reset_index()
                .sort_values("Avg_Confidence")
                .rename(columns={"Com_Name": "Species", "UK_Status": "Status",
                                 "Avg_Confidence": "Avg Confidence"})
            )
            summary["Avg Confidence"] = summary["Avg Confidence"].round(3)
            st.dataframe(summary, hide_index=True)

    st.divider()

    # ── Review Recording: Top Species to Check + Confidence by Hour ──
    st.subheader("Review Recording: Top Species to Check")

    if len(review_df) == 0:
        st.info("No 'Review Recording' rows in the current filter.")
    else:
        top_review = (
            review_df["Sci_Name"]
            .value_counts()
            .head(20)
            .reset_index()
        )
        top_review.columns = ["Sci_Name", "Count"]
        top_review = top_review.sort_values("Count", ascending=True)

        fig = px.bar(
            top_review,
            x="Count", y="Sci_Name", orientation="h",
            title="Top 20 Latin Names Needing Review",
            color="Count",
            color_continuous_scale=[[0, "#c4a07a"], [1, "#7a5c3d"]],
            labels={"Count": "Detections", "Sci_Name": ""},
        )
        fig.update_coloraxes(showscale=False)
        fig.update_traces(marker_line_width=0)
        st.plotly_chart(style_fig(fig), width="stretch")

        st.subheader("Review Recording: Confidence by Hour")

        conf_hour = (
            review_df.groupby("hour")["Confidence"]
            .mean()
            .reset_index(name="Avg_Confidence")
        )
        fig = px.line(
            conf_hour,
            x="hour", y="Avg_Confidence",
            markers=True,
            title="Average Confidence by Hour (Review Recording only)",
            labels={"hour": "Hour of day", "Avg_Confidence": "Avg confidence"},
        )
        fig.update_traces(
            line=dict(color=TERTIARY, width=2),
            marker=dict(size=5, color=TERTIARY),
        )
        fig.update_layout(xaxis=dict(dtick=1))
        st.plotly_chart(style_fig(fig), width="stretch")

    st.divider()

    # ── Validate Review Recording species ──
    st.subheader("Validate a 'Review Recording' Species")

    has_token = False
    try:
        _gh_token = st.secrets["GITHUB_TOKEN"]
        has_token = bool(_gh_token)
    except (KeyError, FileNotFoundError):
        pass

    if not has_token:
        st.info(
            "To validate species from here, configure a `GITHUB_TOKEN` secret "
            "with Contents write permission on the repo."
        )
    elif len(review_df) == 0:
        st.success("No species currently need review!")
    else:
        review_species = (
            review_df[["Sci_Name", "Com_Name"]]
            .drop_duplicates()
            .sort_values("Sci_Name")
        )
        display_labels = (
            review_species["Sci_Name"] + "  (" + review_species["Com_Name"] + ")"
        ).tolist()

        VALID_STATUSES = [
            "Resident", "Summer visitor", "Winter visitor",
            "Passage migrant", "Scarce visitor", "Rare vagrant",
            "Introduced species", "Reintroduced", "Extinct", "False Positive", "Other",
        ]

        with st.form("validate_review_species"):
            chosen = st.selectbox("Species to validate", display_labels)
            new_status = st.selectbox("Assign status", VALID_STATUSES)
            submitted = st.form_submit_button("Save & push to GitHub")

        if submitted:
            idx = display_labels.index(chosen)
            sci_name = review_species.iloc[idx]["Sci_Name"]
            com_name = review_species.iloc[idx]["Com_Name"]

            EXCEL_PATH = "UK_Birds_Generalized_Status.xlsx"
            REPO = "emjgood1995/bird-dashboard"
            TOKEN = st.secrets["GITHUB_TOKEN"]

            # 1. Update the local Excel file
            wb = openpyxl.load_workbook(EXCEL_PATH)
            ws = wb.active

            # Check if species already exists (match on Latin Name in column B)
            existing_row = None
            for row in ws.iter_rows(min_row=2):
                if row[1].value == sci_name:
                    existing_row = row
                    break

            if existing_row is not None:
                existing_row[0].value = com_name
                existing_row[2].value = new_status
            else:
                ws.append([com_name, sci_name, new_status])
            wb.save(EXCEL_PATH)

            # 2. Push to GitHub via Contents API
            api_url = f"https://api.github.com/repos/{REPO}/contents/{EXCEL_PATH}"
            headers = {
                "Authorization": f"Bearer {TOKEN}",
                "Accept": "application/vnd.github+json",
            }

            # GET current SHA
            get_resp = requests.get(api_url, headers=headers, timeout=15)
            if get_resp.status_code != 200:
                st.error(f"GitHub GET failed ({get_resp.status_code}): {get_resp.text}")
            else:
                sha = get_resp.json()["sha"]
                file_bytes = pathlib.Path(EXCEL_PATH).read_bytes()
                encoded = base64.b64encode(file_bytes).decode()

                action = "Update" if existing_row is not None else "Add"
                put_resp = requests.put(
                    api_url,
                    headers=headers,
                    json={
                        "message": f"{action} species status: {sci_name} -> {new_status}",
                        "content": encoded,
                        "sha": sha,
                    },
                    timeout=30,
                )
                if put_resp.status_code in (200, 201):
                    st.cache_data.clear()
                    st.success(
                        f"Saved **{sci_name}** as *{new_status}* and pushed to GitHub."
                    )
                else:
                    st.error(
                        f"GitHub PUT failed ({put_resp.status_code}): {put_resp.text}"
                    )

# ── Records ───────────────────────────────────────────────────────────────
elif page == "Records":

    # ── First & Last Detection per Species ──
    st.subheader("First & Last Detection per Species")
    st.caption("Earliest and most recent date each species was recorded.")

    pr_df = filtered.dropna(subset=["timestamp"]).copy()
    pr_df["year"] = pr_df["timestamp"].dt.year

    pr_years_avail = sorted(pr_df["year"].dropna().unique())
    pr_years = st.multiselect(
        "Filter to years", pr_years_avail,
        default=pr_years_avail, key="pr_years",
    )

    if pr_years:
        pr_df = pr_df[pr_df["year"].isin(pr_years)].copy()

    if len(pr_df) == 0:
        st.info("No data available.")
    else:
        det_range = pr_df.groupby("Com_Name")["timestamp"].agg(["min", "max"]).reset_index()
        det_range.columns = ["Species", "Earliest_Detection", "Latest_Detection"]
        det_range["First Detected"] = det_range["Earliest_Detection"].dt.strftime("%Y-%m-%d")
        det_range["Last Detected"] = det_range["Latest_Detection"].dt.strftime("%Y-%m-%d")
        det_counts = pr_df["Com_Name"].value_counts().rename("Total Detections")
        det_range = det_range.merge(det_counts, left_on="Species", right_index=True)

        pr_k1, pr_k2 = st.columns(2)
        pr_k1.metric("Total species recorded", det_range["Species"].nunique())
        pr_k2.metric("Date range", f"{pr_df['timestamp'].min().strftime('%Y-%m-%d')} to {pr_df['timestamp'].max().strftime('%Y-%m-%d')}")

        st.dataframe(
            det_range[["Species", "First Detected", "Last Detected", "Total Detections"]]
            .sort_values("First Detected"),
            hide_index=True,
        )

        # ── Arrival & Departure Timeline (Gantt) ──
        st.divider()
        st.subheader("Arrival & Departure Timeline")
        st.caption("Presence window for each species — first to last detection.")

        gantt_view = st.radio(
            "View", ["All years combined", "Average across years", "Year-over-year"],
            horizontal=True, key="gantt_view",
        )
        gantt_top_n = st.slider("Top N species", 5, 50, 30, key="gantt_top_n")

        def _gantt_chart(df, x_start, x_end, y, color, color_map=None, color_seq=None, labels=None):
            """Build a Gantt chart with clean hover text (avoids px.timeline 'undefined' bug)."""
            df = df.copy()
            df["_start_str"] = df[x_start].dt.strftime("%Y-%m-%d")
            df["_end_str"] = df[x_end].dt.strftime("%Y-%m-%d")
            fig = px.timeline(
                df, x_start=x_start, x_end=x_end, y=y, color=color,
                color_discrete_map=color_map, color_discrete_sequence=color_seq,
                labels=labels or {},
            )
            # Replace default hover with clean template
            for trace in fig.data:
                trace.hovertemplate = (
                    "<b>%{y}</b><br>"
                    "%{customdata[0]} → %{customdata[1]}"
                    "<extra>%{fullData.name}</extra>"
                )
                # Build customdata from the matching rows
                mask = df[y] == trace.name if y == color else df[color].astype(str) == trace.name
                if mask.any():
                    trace.customdata = df.loc[mask, ["_start_str", "_end_str"]].values
            # Scale height to number of rows so bars don't get clipped
            n_rows = df[y].nunique()
            fig.update_layout(height=max(400, n_rows * 22))
            return fig

        if gantt_view == "All years combined":
            gantt_df = det_range.nlargest(gantt_top_n, "Total Detections").copy()
            # Ensure single-day detections are visible
            mask = gantt_df["Earliest_Detection"] == gantt_df["Latest_Detection"]
            gantt_df.loc[mask, "Latest_Detection"] = gantt_df.loc[mask, "Latest_Detection"] + pd.Timedelta(days=1)
            # Merge UK_Status for colouring
            sp_status = pr_df.drop_duplicates("Com_Name")[["Com_Name", "UK_Status"]]
            gantt_df = gantt_df.merge(sp_status, left_on="Species", right_on="Com_Name", how="left")
            gantt_df["UK_Status"] = gantt_df["UK_Status"].fillna("Review Recording")
            gantt_df = gantt_df.sort_values("Earliest_Detection")
            cmap = status_color_map(gantt_df["UK_Status"].unique())
            fig = _gantt_chart(gantt_df, "Earliest_Detection", "Latest_Detection", "Species", "UK_Status", color_map=cmap)
            fig.update_yaxes(categoryorder="array", categoryarray=gantt_df["Species"].tolist())
            fig.update_layout(yaxis_title="", xaxis_title="")
            st.plotly_chart(style_fig(fig), width="stretch")

        elif gantt_view == "Average across years":
            # Average arrival/departure day-of-year across years, projected onto a reference year
            per_yr = pr_df.groupby(["Com_Name", "year"])["timestamp"].agg(["min", "max"]).reset_index()
            per_yr.columns = ["Species", "year", "first", "last"]
            per_yr["first_doy"] = per_yr["first"].dt.dayofyear
            per_yr["last_doy"] = per_yr["last"].dt.dayofyear
            avg_doy = per_yr.groupby("Species")[["first_doy", "last_doy"]].mean().reset_index()
            avg_doy["n_years"] = per_yr.groupby("Species")["year"].nunique().values
            # Top N by total detections
            top_species = pr_df["Com_Name"].value_counts().head(gantt_top_n).index.tolist()
            avg_doy = avg_doy[avg_doy["Species"].isin(top_species)].copy()
            # Project onto reference year 2000 (leap year, so all 366 days valid)
            avg_doy["Start"] = pd.to_datetime("2000-01-01") + pd.to_timedelta(avg_doy["first_doy"].round() - 1, unit="D")
            avg_doy["End"] = pd.to_datetime("2000-01-01") + pd.to_timedelta(avg_doy["last_doy"].round() - 1, unit="D")
            # Pad single-day
            mask = avg_doy["Start"] == avg_doy["End"]
            avg_doy.loc[mask, "End"] = avg_doy.loc[mask, "End"] + pd.Timedelta(days=1)
            # Merge UK_Status
            sp_status = pr_df.drop_duplicates("Com_Name")[["Com_Name", "UK_Status"]]
            avg_doy = avg_doy.merge(sp_status, left_on="Species", right_on="Com_Name", how="left")
            avg_doy["UK_Status"] = avg_doy["UK_Status"].fillna("Review Recording")
            avg_doy = avg_doy.sort_values("first_doy")
            cmap = status_color_map(avg_doy["UK_Status"].unique())
            fig = _gantt_chart(avg_doy, "Start", "End", "Species", "UK_Status", color_map=cmap)
            fig.update_yaxes(categoryorder="array", categoryarray=avg_doy["Species"].tolist())
            # Format x-axis as month names (reference year)
            fig.update_layout(
                yaxis_title="", xaxis_title="",
                xaxis=dict(
                    tickformat="%b",
                    dtick="M1",
                ),
            )
            st.plotly_chart(style_fig(fig), width="stretch")

        else:
            # Year-over-year mode
            yoy = pr_df.groupby(["Com_Name", "year"])["timestamp"].agg(["min", "max"]).reset_index()
            yoy.columns = ["Species", "year", "Start", "End"]
            # Pad single-day
            mask = yoy["Start"] == yoy["End"]
            yoy.loc[mask, "End"] = yoy.loc[mask, "End"] + pd.Timedelta(days=1)
            # Top N by total detections
            top_species = pr_df["Com_Name"].value_counts().head(gantt_top_n).index.tolist()
            yoy = yoy[yoy["Species"].isin(top_species)].copy()
            yoy["year_str"] = yoy["year"].astype(str)
            yoy["Label"] = yoy["Species"] + " (" + yoy["year_str"] + ")"
            yoy = yoy.sort_values(["Species", "year"])
            fig = _gantt_chart(yoy, "Start", "End", "Label", "year_str", color_seq=NATURE_PALETTE, labels={"year_str": "Year"})
            fig.update_yaxes(categoryorder="array", categoryarray=yoy["Label"].tolist())
            fig.update_layout(yaxis_title="", xaxis_title="")
            st.plotly_chart(style_fig(fig), width="stretch")

    st.divider()

    # ── Rarest Visitors ──
    st.subheader("Rarest Visitors")
    st.caption("Species with the fewest total detections — when they appeared and at what confidence.")
    pr_rarest_n = st.slider("N rarest species", 5, 30, 15, key="pr_rarest_n")

    if len(pr_df) == 0:
        st.info("No data available.")
    else:
        # Identify the N rarest species
        species_counts = pr_df["Com_Name"].value_counts()
        rare_species = species_counts.tail(pr_rarest_n).index.tolist()
        rare_df = pr_df[pr_df["Com_Name"].isin(rare_species)].copy()
        rare_df["date"] = rare_df["timestamp"].dt.date

        # Timeline scatter — coloured by UK status
        cmap = status_color_map(rare_df["UK_Status"].dropna().unique())
        # Order species by detection count (fewest at top)
        species_order = species_counts.loc[rare_species].sort_values().index.tolist()

        fig = px.scatter(
            rare_df, x="date", y="Com_Name",
            color="UK_Status",
            size="Confidence",
            size_max=12,
            title=f"Rarest {pr_rarest_n} Species — Detection Timeline",
            labels={"date": "Date", "Com_Name": "Species", "UK_Status": "UK Status",
                    "Confidence": "Confidence"},
            color_discrete_map=cmap,
            category_orders={"Com_Name": species_order},
            hover_data={"Confidence": ":.2f", "date": True},
        )
        fig.update_traces(marker=dict(opacity=0.8, line=dict(width=0.5, color="#1a2416")))
        fig.update_layout(height=max(400, len(species_order) * 22))
        st.plotly_chart(style_fig(fig), width="stretch")

        # Detail table
        rare_table = (
            rare_df.groupby("Com_Name")
            .agg(
                Detections=("Com_Name", "size"),
                Avg_Confidence=("Confidence", "mean"),
                First_Seen=("timestamp", "min"),
                Last_Seen=("timestamp", "max"),
                UK_Status=("UK_Status", "first"),
            )
            .reset_index()
        )
        rare_table["First Seen"] = rare_table["First_Seen"].dt.strftime("%Y-%m-%d")
        rare_table["Last Seen"] = rare_table["Last_Seen"].dt.strftime("%Y-%m-%d")
        rare_table["Avg Confidence"] = rare_table["Avg_Confidence"].round(3)
        rare_table = rare_table.sort_values("Detections")

        st.dataframe(
            rare_table[["Com_Name", "Detections", "Avg Confidence", "First Seen", "Last Seen", "UK_Status"]]
            .rename(columns={"Com_Name": "Species", "UK_Status": "UK Status"}),
            hide_index=True,
        )

    st.divider()

    # ── Longest Detection Streak ──
    st.subheader("Longest Detection Streak")

    if len(pr_df) == 0:
        st.info("No data available.")
    else:
        def longest_streak(dates):
            """Compute longest run of consecutive days."""
            if len(dates) == 0:
                return 0
            unique_days = sorted(set(dates))
            best = 1
            current = 1
            for i in range(1, len(unique_days)):
                if (unique_days[i] - unique_days[i - 1]).days == 1:
                    current += 1
                    best = max(best, current)
                else:
                    current = 1
            return best

        pr_df["det_date"] = pr_df["timestamp"].dt.date
        streak_data = (
            pr_df.groupby("Com_Name")["det_date"]
            .apply(lambda x: longest_streak(x.tolist()))
            .reset_index(name="Longest_Streak")
            .sort_values("Longest_Streak", ascending=False)
        )

        st.metric("Top streak", f"{streak_data['Longest_Streak'].max()} days" if len(streak_data) else "—")
        st.dataframe(
            streak_data.rename(columns={"Com_Name": "Species", "Longest_Streak": "Longest Streak (days)"}),
            hide_index=True,
        )

    # ── Phenology Calendar ──
    st.divider()
    st.subheader("Phenology Calendar")
    st.caption("When each species is active throughout the year.")

    pheno_col1, pheno_col2 = st.columns(2)
    with pheno_col1:
        pheno_top_n = st.slider("Top N species", 5, 40, 25, key="pheno_top_n")
    with pheno_col2:
        pheno_metric = st.radio("Metric", ["Detections", "Days active"], horizontal=True, key="pheno_metric")

    if len(filtered) > 0:
        if pheno_metric == "Detections":
            pheno_agg = filtered.groupby(["Com_Name", "month"]).size().reset_index(name="value")
        else:
            pheno_agg = filtered.groupby(["Com_Name", "month"])["Date"].nunique().reset_index(name="value")

        pheno_totals = pheno_agg.groupby("Com_Name")["value"].sum().nlargest(pheno_top_n)
        pheno_species = pheno_totals.index.tolist()
        pheno_agg = pheno_agg[pheno_agg["Com_Name"].isin(pheno_species)]

        pheno_pivot = pheno_agg.pivot(index="Com_Name", columns="month", values="value").fillna(0)
        pheno_pivot = pheno_pivot.reindex(columns=range(1, 13), fill_value=0)
        pheno_pivot = pheno_pivot.loc[pheno_species]  # sort by total descending

        fig = px.imshow(
            pheno_pivot.values,
            x=list(MONTH_LABELS.values()),
            y=pheno_pivot.index.tolist(),
            title=f"Phenology Calendar · {pheno_metric}",
            color_continuous_scale=HEATMAP_SCALE,
            labels={"x": "Month", "y": "Species", "color": pheno_metric},
            aspect="auto",
        )
        fig.update_layout(
            xaxis=dict(dtick=1),
            yaxis=dict(dtick=1),
            height=max(400, len(pheno_pivot) * 22),
            coloraxis_colorbar=dict(
                title=pheno_metric,
                tickfont=dict(size=11, color="#4a5c44"),
                title_font=dict(size=12, color="#4a5c44"),
                thickness=14,
            ),
        )
        st.plotly_chart(style_fig(fig), width="stretch")
    else:
        st.info("No data available for the current filters.")

    # ── Classify Unclassified Species ──
    st.divider()
    st.subheader("Classify Unclassified Species")
    st.caption("Species not yet assigned a diet category.")

    unclassified = df[df["Diet"] == "Unclassified"][["Sci_Name", "Com_Name"]].drop_duplicates().sort_values("Sci_Name")

    if len(unclassified) == 0:
        st.success("All species have been classified!")
    else:
        st.warning(f"{len(unclassified)} species need diet classification.")
        labels = (unclassified["Sci_Name"] + "  (" + unclassified["Com_Name"] + ")").tolist()

        DIET_CATEGORIES = ["Insectivore", "Granivore", "Omnivore", "Frugivore",
                           "Carnivore", "Piscivore", "Herbivore"]

        with st.form("classify_diet"):
            chosen = st.selectbox("Species", labels)
            diet = st.selectbox("Diet category", DIET_CATEGORIES)
            submitted = st.form_submit_button("Save classification")

        if submitted:
            idx = labels.index(chosen)
            sci_name = unclassified.iloc[idx]["Sci_Name"]

            diet_data = load_diet_map()
            diet_data[sci_name] = diet
            with open("species_diet.json", "w") as f:
                json.dump(diet_data, f, indent=2, sort_keys=True)

            st.cache_data.clear()
            st.success(f"Classified {sci_name} as {diet}.")
            st.rerun()

# ── Nearby Sightings ─────────────────────────────────────────────────────
elif page == "Nearby Sightings":
    st.subheader("Nearby Sightings")
    st.caption("Recent bird observations from iNaturalist near your recording station.")

    # Extract station location
    if "Lat" not in df.columns or "Lon" not in df.columns or df["Lat"].dropna().empty:
        st.warning("No location data available for your station.")
    else:
        stn_lat = 52.2387
        stn_lon = 0.2477

        col_r, col_d, col_reset = st.columns([1, 1, 0.5])
        with col_r:
            radius_km = st.slider("Radius (km)", 5, 50, 25)
        with col_d:
            days_back = st.slider("Days back", 7, 90, 30)
        with col_reset:
            st.markdown("<br>", unsafe_allow_html=True)
            if st.button("Reset map"):
                st.session_state["inat_map_key"] = st.session_state.get("inat_map_key", 0) + 1

        map_key = st.session_state.get("inat_map_key", 0)

        data = fetch_inat_nearby(stn_lat, stn_lon, radius_km, days_back)

        if data is None:
            st.warning("Could not reach iNaturalist. Please try again later.")
        else:
            results = data.get("results", [])
            rows = []
            for obs in results:
                taxon = obs.get("taxon") or {}
                loc = obs.get("location", "")
                if not loc or not taxon.get("name"):
                    continue
                olat, olng = loc.split(",")
                rows.append({
                    "species": taxon.get("preferred_common_name", taxon.get("name", "")),
                    "sci_name": taxon.get("name", ""),
                    "lat": float(olat),
                    "lon": float(olng),
                    "observed_on": obs.get("observed_on", ""),
                    "place_guess": obs.get("place_guess", ""),
                    "uri": obs.get("uri", ""),
                })

            if not rows:
                st.info("No recent bird observations found nearby.")
            else:
                inat_df = pd.DataFrame(rows)

                # Cross-reference with garden detections
                garden_sci = set(df["Sci_Name"].dropna().unique())
                inat_df["Seen in garden"] = inat_df["sci_name"].isin(garden_sci)

                event = st.pydeck_chart(pdk.Deck(
                    map_style="https://basemaps.cartocdn.com/gl/positron-gl-style/style.json",
                    initial_view_state=pdk.ViewState(
                        latitude=stn_lat, longitude=stn_lon, zoom=10, pitch=0,
                    ),
                    layers=[
                        pdk.Layer(
                            "ScatterplotLayer",
                            id="observations",
                            data=inat_df,
                            get_position=["lon", "lat"],
                            get_radius=350,
                            get_fill_color=[234, 85, 50, 210],
                            get_line_color=[255, 255, 255, 220],
                            line_width_min_pixels=1,
                            stroked=True,
                            pickable=True,
                            auto_highlight=True,
                            highlight_color=[255, 200, 0, 200],
                        ),
                        pdk.Layer(
                            "ScatterplotLayer",
                            id="station",
                            data=pd.DataFrame([{"lat": stn_lat, "lon": stn_lon, "species": "Your station"}]),
                            get_position=["lon", "lat"],
                            get_radius=500,
                            get_fill_color=[30, 100, 220, 230],
                            get_line_color=[255, 255, 255, 255],
                            line_width_min_pixels=2,
                            stroked=True,
                            pickable=True,
                        ),
                    ],
                    tooltip={
                        "html": "<div style='padding:6px 10px'>"
                                "<b style='font-size:15px'>{species}</b><br/>"
                                "<i style='color:#555'>{sci_name}</i><br/>"
                                "<span style='color:#333'>📅 {observed_on}</span><br/>"
                                "<span style='color:#333'>📍 {place_guess}</span>"
                                "</div>",
                        "style": {"backgroundColor": "white",
                                  "color": "#222", "fontSize": "14px",
                                  "borderRadius": "8px",
                                  "boxShadow": "0 2px 8px rgba(0,0,0,0.15)"},
                    },
                ), on_select="rerun", selection_mode="multi-object", key=f"inat_map_{map_key}", height=600)

                st.divider()

                # Filter table to selection if points clicked on map
                selected_indices = []
                if event and event.selection and event.selection.get("indices"):
                    # indices is dict keyed by layer id
                    selected_indices = event.selection["indices"].get("observations", [])

                if selected_indices:
                    table_df = inat_df.iloc[selected_indices]
                    st.caption(f"Showing {len(selected_indices)} selected observation(s). Click empty area on map to reset.")
                else:
                    table_df = inat_df

                total_obs = len(table_df)
                unique_species = table_df["sci_name"].nunique()
                garden_overlap = table_df.loc[table_df["Seen in garden"], "sci_name"].nunique()
                most_reported = table_df["species"].value_counts().idxmax()

                m1, m2, m3, m4 = st.columns(4)
                m1.metric("Observations", total_obs)
                m2.metric("Unique species", unique_species)
                m3.metric("Also in your garden", garden_overlap)
                m4.metric("Most reported", most_reported)

                display_df = table_df[["species", "sci_name", "observed_on", "place_guess", "Seen in garden", "uri"]].copy()
                display_df.columns = ["Species", "Scientific name", "Date", "Location", "Seen in garden", "iNat link"]
                display_df = display_df.sort_values("Date", ascending=False).reset_index(drop=True)

                st.dataframe(
                    display_df,
                    column_config={
                        "iNat link": st.column_config.LinkColumn("iNat link", display_text="View"),
                    },
                    use_container_width=True,
                )

# ── Species Explorer ──────────────────────────────────────────────────────
elif page == "Species Explorer":
    st.subheader("Species Explorer")

    se_species_pairs = (
        filtered[["Com_Name", "Sci_Name"]]
        .dropna()
        .drop_duplicates()
        .sort_values("Com_Name")
        .reset_index(drop=True)
    )

    if len(se_species_pairs) == 0:
        st.info("No species available for the current filters.")
    else:
        se_labels = (se_species_pairs["Com_Name"] + "  (" + se_species_pairs["Sci_Name"] + ")").tolist()
        today_str = str(pd.Timestamp.now().date())
        bird_of_day_idx = int(hashlib.md5(today_str.encode()).hexdigest(), 16) % len(se_labels)

        chosen_label = st.selectbox("Choose a species", se_labels, index=bird_of_day_idx, key="se_species")
        chosen_idx = se_labels.index(chosen_label)
        se_com = se_species_pairs.iloc[chosen_idx]["Com_Name"]
        se_sci = se_species_pairs.iloc[chosen_idx]["Sci_Name"]

        # Fetch Wikipedia info — try scientific name first, fall back to common name
        wiki = fetch_wiki_summary(se_sci)
        if wiki is None:
            wiki = fetch_wiki_summary(se_com)

        if wiki is not None:
            img_col, text_col = st.columns([1, 2], gap="large")
            with img_col:
                if wiki["thumbnail_url"]:
                    st.image(wiki["thumbnail_url"], use_container_width=True)
                else:
                    st.info("No image available.")
            with text_col:
                st.markdown(f"### {se_com}")
                st.markdown(f"*{se_sci}*")

                _se_status = filtered.loc[filtered["Sci_Name"] == se_sci, "UK_Status"].mode()
                _se_diet = filtered.loc[filtered["Sci_Name"] == se_sci, "Diet"].mode()
                _se_status_val = _se_status.iloc[0] if len(_se_status) else "Unknown"
                _se_diet_val = _se_diet.iloc[0] if len(_se_diet) else "Unclassified"
                _stat_color = STATUS_COLORS.get(_se_status_val, "#8c9c8c")
                _diet_color = DIET_COLORS.get(_se_diet_val, "#8c9c8c")
                st.markdown(
                    f'<span style="background:{_stat_color};color:#fff;padding:3px 10px;border-radius:8px;font-size:0.85rem;font-weight:600;margin-right:8px">{_se_status_val}</span>'
                    f'<span style="background:{_diet_color};color:#fff;padding:3px 10px;border-radius:8px;font-size:0.85rem;font-weight:600">{_se_diet_val}</span>',
                    unsafe_allow_html=True,
                )

                st.markdown(wiki["extract"])
                if wiki["page_url"]:
                    st.markdown(f"[Read more on Wikipedia]({wiki['page_url']})")

            if chosen_idx == bird_of_day_idx:
                st.caption("Bird of the day — changes daily, seeded by today's date.")

            first_sentence = wiki["extract"].split(". ")[0]
            if first_sentence:
                st.info(f"Fun fact: {first_sentence}.")
        else:
            st.markdown(f"### {se_com}")
            st.markdown(f"*{se_sci}*")
            _se_status = filtered.loc[filtered["Sci_Name"] == se_sci, "UK_Status"].mode()
            _se_diet = filtered.loc[filtered["Sci_Name"] == se_sci, "Diet"].mode()
            _se_status_val = _se_status.iloc[0] if len(_se_status) else "Unknown"
            _se_diet_val = _se_diet.iloc[0] if len(_se_diet) else "Unclassified"
            _stat_color = STATUS_COLORS.get(_se_status_val, "#8c9c8c")
            _diet_color = DIET_COLORS.get(_se_diet_val, "#8c9c8c")
            st.markdown(
                f'<span style="background:{_stat_color};color:#fff;padding:3px 10px;border-radius:8px;font-size:0.85rem;font-weight:600;margin-right:8px">{_se_status_val}</span>'
                f'<span style="background:{_diet_color};color:#fff;padding:3px 10px;border-radius:8px;font-size:0.85rem;font-weight:600">{_se_diet_val}</span>',
                unsafe_allow_html=True,
            )
            st.warning("Could not fetch information from Wikipedia.")

        # Bird song from Wikimedia Commons
        bird_audio = fetch_bird_audio(se_sci)
        if bird_audio and bird_audio["file"]:
            st.markdown("**Listen**")
            st.audio(bird_audio["file"], format=bird_audio["format"])
            st.caption(
                f"Source: [Wikimedia Commons]({bird_audio['page']})"
            )

        # Detection summary for selected species
        st.divider()
        st.markdown(f"#### Detection Summary: {se_com}")
        sp_df = filtered[filtered["Com_Name"] == se_com].copy()

        if len(sp_df) == 0:
            st.info("No detections for this species in the current filters.")
        else:
            sk1, sk2, sk3, sk4 = st.columns(4)
            sk1.metric("Total Detections", f"{len(sp_df):,}")
            sp_ts = sp_df.dropna(subset=["timestamp"])
            if len(sp_ts) > 0:
                sk2.metric("Last Seen", sp_ts['timestamp'].max().strftime('%Y-%m-%d'))
                peak_hour = sp_ts["hour"].value_counts().idxmax()
                sk3.metric("Peak Hour", f"{peak_hour}:00")
                peak_month = sp_ts["month"].value_counts().idxmax()
                sk4.metric("Peak Month", MONTH_LABELS.get(peak_month, str(peak_month)))

        # Update species status & diet form
        st.divider()
        current_status = sp_df["UK_Status"].mode().iloc[0] if len(sp_df) > 0 else "Unknown"
        current_diet = sp_df["Diet"].mode().iloc[0] if len(sp_df) > 0 else "Unclassified"
        st.markdown(f"#### Update Status & Diet for {se_com}")
        st.caption(f"Current status: **{current_status}** · Current diet: **{current_diet}**")

        has_token = False
        try:
            _gh_token = st.secrets["GITHUB_TOKEN"]
            has_token = bool(_gh_token)
        except (KeyError, FileNotFoundError):
            pass

        if not has_token:
            st.info(
                "To update species status, configure a `GITHUB_TOKEN` secret "
                "with Contents write permission on the repo."
            )
        else:
            SE_STATUSES = [
                "Resident", "Summer visitor", "Winter visitor",
                "Passage migrant", "Scarce visitor", "Rare vagrant",
                "Introduced species", "Reintroduced", "Extinct", "False Positive", "Other",
            ]
            SE_DIETS = ["Insectivore", "Granivore", "Omnivore", "Frugivore",
                        "Carnivore", "Piscivore", "Herbivore"]
            default_status_idx = SE_STATUSES.index(current_status) if current_status in SE_STATUSES else 0
            default_diet_idx = SE_DIETS.index(current_diet) if current_diet in SE_DIETS else 0

            with st.form("se_update_status"):
                se_new_status = st.selectbox("Assign status", SE_STATUSES, index=default_status_idx, key="se_new_status")
                se_new_diet = st.selectbox("Assign diet", SE_DIETS, index=default_diet_idx, key="se_new_diet")
                se_submitted = st.form_submit_button("Save & push to GitHub")

            if se_submitted:
                # ── Save diet to local JSON ──
                diet_changed = se_new_diet != current_diet
                if diet_changed:
                    diet_data = load_diet_map()
                    diet_data[se_sci] = se_new_diet
                    with open("species_diet.json", "w") as f:
                        json.dump(diet_data, f, indent=2, sort_keys=True)

                # ── Save status to Excel & push to GitHub ──
                EXCEL_PATH = "UK_Birds_Generalized_Status.xlsx"
                REPO = "emjgood1995/bird-dashboard"
                TOKEN = st.secrets["GITHUB_TOKEN"]

                wb = openpyxl.load_workbook(EXCEL_PATH)
                ws = wb.active

                # Check if species already exists (match on Latin Name in column B)
                existing_row = None
                for row in ws.iter_rows(min_row=2):
                    if row[1].value == se_sci:
                        existing_row = row
                        break

                if existing_row is not None:
                    existing_row[0].value = se_com
                    existing_row[2].value = se_new_status
                else:
                    ws.append([se_com, se_sci, se_new_status])
                wb.save(EXCEL_PATH)

                api_url = f"https://api.github.com/repos/{REPO}/contents/{EXCEL_PATH}"
                headers = {
                    "Authorization": f"Bearer {TOKEN}",
                    "Accept": "application/vnd.github+json",
                }

                get_resp = requests.get(api_url, headers=headers, timeout=15)
                if get_resp.status_code != 200:
                    st.error(f"GitHub GET failed ({get_resp.status_code}): {get_resp.text}")
                else:
                    sha = get_resp.json()["sha"]
                    file_bytes = pathlib.Path(EXCEL_PATH).read_bytes()
                    encoded = base64.b64encode(file_bytes).decode()

                    action = "Update" if existing_row is not None else "Add"
                    put_resp = requests.put(
                        api_url,
                        headers=headers,
                        json={
                            "message": f"{action} species status: {se_sci} -> {se_new_status}",
                            "content": encoded,
                            "sha": sha,
                        },
                        timeout=30,
                    )
                    if put_resp.status_code in (200, 201):
                        st.cache_data.clear()
                        parts = [f"Status: *{se_new_status}*"]
                        if diet_changed:
                            parts.append(f"Diet: *{se_new_diet}*")
                        st.success(f"Saved **{se_sci}** — {', '.join(parts)}. Pushed to GitHub.")
                    else:
                        st.error(f"GitHub PUT failed ({put_resp.status_code}): {put_resp.text}")

# ── Birthday Easter Egg ──────────────────────────────────────────────────────
elif page == "\U0001f382":
    if "bday_balloons_shown" not in st.session_state:
        st.session_state["bday_balloons_shown"] = True
        st.balloons()

    st.markdown(
        '<div class="birthday-banner">&#x2727; H A P P Y &nbsp; B I R T H D A Y &#x2727;</div>',
        unsafe_allow_html=True,
    )

    # Filter to Feb 23 across all years
    bday_all = filtered[
        (filtered["timestamp"].dt.month == 2) & (filtered["timestamp"].dt.day == 23)
    ]

    if bday_all.empty:
        st.info("No birds recorded on your birthday yet — check back tonight!")
    else:
        # Year selector
        bday_years = sorted(bday_all["timestamp"].dt.year.unique())
        bday_year_opts = ["All years"] + [str(y) for y in bday_years]
        bday_year_sel = st.selectbox("Year", bday_year_opts, index=0, key="bday_year")

        if bday_year_sel == "All years":
            bday = bday_all.copy()
            bday_year_label = "All Years"
        else:
            bday = bday_all[bday_all["timestamp"].dt.year == int(bday_year_sel)].copy()
            bday_year_label = bday_year_sel

        # KPI metrics
        k1, k2, k3 = st.columns(3)
        k1.metric("Total Detections", f"{len(bday):,}")
        k2.metric("Unique Species", f"{bday['Com_Name'].nunique():,}")
        k3.metric("Years of Data", f"{bday['timestamp'].dt.year.nunique()}")

        st.divider()

        # Birthday Birds — horizontal bar chart
        st.subheader("Birthday Birds")
        bday_sp = (
            bday["Com_Name"].value_counts()
            .reset_index()
        )
        bday_sp.columns = ["Species", "Count"]
        bday_sp = bday_sp.sort_values("Count", ascending=True)

        fig_sp = px.bar(
            bday_sp, x="Count", y="Species", orientation="h",
            title=f"Species Detected on Feb 23 · {bday_year_label}",
            color="Count",
            color_continuous_scale=[[0, NATURE_PALETTE[6]], [1, NATURE_PALETTE[0]]],
            labels={"Count": "Detections", "Species": ""},
        )
        fig_sp.update_coloraxes(showscale=False)
        fig_sp.update_traces(marker_line_width=0)
        st.plotly_chart(style_fig(fig_sp), width="stretch")

        st.divider()

        # Hourly activity chart (same style as Community page)
        st.subheader("Hourly Activity on Feb 23")
        bday_show_species = st.checkbox("Show by species", value=False, key="bday_by_species")

        if bday_show_species:
            bday_top = bday["Com_Name"].value_counts().head(20).index.tolist()
            bday_tod = bday[bday["Com_Name"].isin(bday_top)].copy()
            bday_sp_hour = bday_tod.groupby(["hour", "Com_Name"]).size().reset_index(name="Count")
            bday_sp_cmap = {
                sp: NATURE_PALETTE[i % len(NATURE_PALETTE)]
                for i, sp in enumerate(bday_top)
            }
            fig_hr = px.area(
                bday_sp_hour, x="hour", y="Count",
                color="Com_Name",
                title=f"Activity by Hour · Feb 23 · {bday_year_label}",
                labels={"hour": "Hour of day", "Count": "Detections", "Com_Name": "Species"},
                category_orders={"Com_Name": bday_top},
                color_discrete_map=bday_sp_cmap,
            )
            fig_hr.update_layout(xaxis=dict(dtick=1))
            fig_hr.update_traces(marker_line_width=0)
            bday_hourly_total = bday.groupby("hour").size().reset_index(name="Count")
            fig_hr.add_scatter(
                x=bday_hourly_total["hour"], y=bday_hourly_total["Count"],
                mode="lines+markers",
                line=dict(color="#1a2416", width=2.5, dash="dot"),
                marker=dict(size=5, color="#1a2416"),
                name="Total", showlegend=True,
            )
        else:
            bday_hourly = bday.groupby("hour").size().reset_index(name="Count")
            fig_hr = px.area(
                bday_hourly, x="hour", y="Count",
                title=f"Activity by Hour · Feb 23 · {bday_year_label}",
                labels={"hour": "Hour of day", "Count": "Detections"},
            )
            fig_hr.update_traces(
                line=dict(color=PRIMARY, width=2),
                fillcolor="rgba(61,107,68,0.14)",
                marker=dict(size=5, color=PRIMARY),
                mode="lines+markers",
            )
            fig_hr.update_layout(xaxis=dict(dtick=1))
        st.plotly_chart(style_fig(fig_hr), width="stretch")

        st.divider()

        # Fun birthday stats
        st.subheader("Fun Birthday Stats")
        f1, f2, f3 = st.columns(3)

        earliest_row = bday.loc[bday["timestamp"].dt.time.idxmin()]
        earliest_time = earliest_row["timestamp"].strftime("%H:%M")
        f1.metric("Earliest Bird", earliest_row["Com_Name"], delta=earliest_time, delta_color="off")

        most_common = bday["Com_Name"].value_counts().idxmax()
        most_common_n = bday["Com_Name"].value_counts().max()
        f2.metric("Most Common Birthday Bird", most_common, delta=f"{most_common_n} detections", delta_color="off")

        rarest = bday["Com_Name"].value_counts()
        rarest_sp = rarest[rarest == rarest.min()]
        rarest_label = rarest_sp.index[0] if len(rarest_sp) == 1 else f"{rarest_sp.index[0]} (+{len(rarest_sp)-1} more)"
        f3.metric("Rarest Birthday Visitor", rarest_label, delta=f"{rarest.min()} detection(s)", delta_color="off")
